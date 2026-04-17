# 子维度与工作关系表生成脚本
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============ Sheet 1: 说明 ============
ws_info = wb.active
ws_info.title = "使用说明"
ws_info['A1'] = '子维度与工作适配关系表 — 使用说明'
ws_info['A1'].font = Font(name='Arial', bold=True, size=13)
ws_info['A3'] = '■ 关系说明'
ws_info['A3'].font = Font(name='Arial', bold=True, size=11)

guide = [
    '① 本表展示27个子维度与50个工作岗位之间的适配关系',
    '② 适配关系基于每个子维度的"成长型"（高分=能力强）和"等级型"（高分=能力弱）两种模式综合判定',
    '③ 成长型模式高分 → 推荐岗位；等级型模式高分 → 不推荐岗位',
    '④ 评分标准：',
    '    · 核心匹配（★★★★★）：该子维度能力强的成长型人群强烈推荐，反之等级型人群强烈不建议',
    '    · 延伸匹配（★★★）：该维度有正向加分，但非决定性因素',
    '    · 不相关（☆）：该维度高低对该工作影响较小',
    '⑤ 注意：同一维度可能既有"成长型高分推荐"也有"等级型高分推荐"（因为不同模式适配不同工作）',
    '    · 例如：共情力成长型→推荐销售；共情力等级型→推荐独立研发',
]
for i, line in enumerate(guide):
    ws_info.cell(row=4+i, column=1, value=line).font = Font(name='Arial', size=10)

ws_info.column_dimensions['A'].width = 90

# ============ Sheet 2: 汇总表 ============
ws_summary = wb.create_sheet("适配关系总表")
ws_summary.sheet_view.showGridLines = False

ws_summary['A1'] = '27个子维度 × 50个岗位 — 适配关系矩阵（成长型推荐视角）'
ws_summary['A1'].font = Font(name='Arial', bold=True, size=12)
ws_summary.merge_cells('A1:BX1')
ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_summary.row_dimensions[1].height = 28

all_jobs = [
    '后端开发', '前端开发', '算法工程', '数据分析', '硬件工程', '测试工程',
    'UI设计', 'UX设计', '品牌设计', '工业设计', '视频剪辑', '插画师',
    '产品经理', '项目经理', '创新研究', '内容创作', '知识产权',
    '战略规划', '管理咨询', '投资分析', '商业分析', '行业研究',
    '电商运营', '用户运营', '内容运营', '供应链', '质量管理',
    '品牌营销', '数字营销', '销售经理', '大客户', '客户成功', 'BD商务',
    '财务分析', '审计', '投融资', '税务',
    'HRBP', '招聘', '培训发展', 'OD组织',
    '行政', '法务', '数据录入', '客服',
    '运营总监', '市场总监', 'CTO', 'HRD', 'CEO创业'
]

all_jobs_full = [
    '后端开发工程师', '前端开发工程师', '算法工程师', '数据分析师', '硬件工程师', '测试工程师',
    'UI设计师', 'UX设计师', '品牌设计师', '工业设计师', '视频剪辑师', '插画师',
    '产品经理', '项目经理', '创新研究员', '内容创作者', '知识产权专员',
    '战略规划师', '管理咨询顾问', '投资分析师', '商业分析师', '行业研究员',
    '电商运营', '用户运营', '内容运营', '供应链专员', '质量管理专员',
    '品牌营销经理', '数字营销专员', '销售经理', '大客户经理', '客户成功经理', 'BD商务拓展',
    '财务分析师', '审计专员', '投融资专员', '税务专员',
    'HRBP', '招聘专员', '培训发展专员', 'OD组织发展专员',
    '行政专员', '法务专员', '数据录入员', '客服专员',
    '运营总监', '市场总监', '技术VP/CTO', 'HRD/CHO', 'CEO/创业者'
]

border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 表头
ws_summary.cell(row=2, column=1, value='子维度').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws_summary.cell(row=2, column=1).fill = PatternFill('solid', fgColor='2F5496')
ws_summary.cell(row=2, column=2, value='大类').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws_summary.cell(row=2, column=2).fill = PatternFill('solid', fgColor='2F5496')
ws_summary.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_summary.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')

for col, job in enumerate(all_jobs, 3):
    c = ws_summary.cell(row=2, column=col, value=job)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws_summary.row_dimensions[2].height = 35

ws_summary.column_dimensions['A'].width = 18
ws_summary.column_dimensions['B'].width = 12
for i in range(3, 53):
    ws_summary.column_dimensions[get_column_letter(i)].width = 9

# 27个子维度数据
dimensions_data = [
    # (名称, 大类, [(岗位索引1-50), 关系级别, 说明]...)
    # 0=后端开发, 1=前端开发, 2=算法, 3=数据, 4=硬件, 5=测试, 6=UI, 7=UX, 8=品牌设计,
    # 9=工业设计, 10=视频剪辑, 11=插画, 12=产品, 13=项目, 14=创新, 15=内容, 16=IP,
    # 17=战略, 18=咨询, 19=投资, 20=商业, 21=行业研究,
    # 22=电商运营, 23=用户运营, 24=内容运营, 25=供应链, 26=质量管理,
    # 27=品牌营销, 28=数字营销, 29=销售, 30=大客户, 31=客户成功, 32=BD,
    # 33=财务分析, 34=审计, 35=投融资, 36=税务,
    # 37=HRBP, 38=招聘, 39=培训, 40=OD,
    # 41=行政, 42=法务, 43=数据录入, 44=客服,
    # 45=运营总监, 46=市场总监, 47=CTO, 48=HRD, 49=CEO创业

    ('注意力/专注力', '认知能力',
     [(0,'H'),(2,'H'),(4,'H'),(5,'H'),(33,'H'),(34,'H'),(43,'H'),
      (1,'M'),(6,'M'),(12,'M'),(22,'M'),(37,'M'),(42,'M'),(23,'M'),
      (29,'L'),(31,'L'),(32,'L'),(41,'L'),(44,'L')]),

    ('思维与洞察力', '认知能力',
     [(12,'H'),(14,'H'),(17,'H'),(18,'H'),(20,'H'),(21,'H'),
      (0,'M'),(2,'M'),(8,'M'),(37,'M'),(40,'M'),
      (43,'L'),(44,'L'),(41,'L')]),

    ('好奇心/学习力', '认知能力',
     [(2,'H'),(14,'H'),(15,'H'),(19,'H'),(21,'H'),(12,'H'),
      (0,'M'),(1,'M'),(3,'M'),(39,'M'),(37,'M'),
      (43,'L'),(44,'L')]),

    ('信息分析/战略思维', '认知能力',
     [(17,'H'),(19,'H'),(21,'H'),(20,'H'),(28,'H'),(33,'H'),
      (12,'M'),(27,'M'),(45,'M'),(37,'M'),(18,'M'),
      (43,'L'),(44,'L'),(41,'L')]),

    ('元认知能力', '认知能力',
     [(12,'H'),(18,'H'),(13,'H'),(45,'H'),(49,'H'),
      (15,'M'),(2,'M'),(39,'M'),(37,'M'),
      (43,'L'),(44,'L')]),

    ('知识体系/跨界迁移', '认知能力',
     [(15,'H'),(12,'H'),(21,'H'),(39,'H'),(49,'H'),
      (18,'M'),(15,'M'),(8,'M'),(37,'M'),
      (43,'L'),(44,'L')]),

    ('开放判断力', '认知能力',
     [(18,'H'),(37,'H'),(40,'H'),(12,'H'),(48,'H'),
      (17,'M'),(27,'M'),(39,'M'),(19,'M'),
      (43,'L'),(44,'L')]),

    ('创造力与创新', '认知能力',
     [(12,'H'),(14,'H'),(6,'H'),(8,'H'),(11,'H'),(15,'H'),
      (1,'M'),(10,'M'),(27,'M'),(28,'M'),
      (34,'L'),(43,'L'),(44,'L')]),

    ('实践落地/反馈迭代', '认知能力',
     [(13,'H'),(45,'H'),(22,'H'),(29,'H'),(12,'H'),
      (0,'M'),(1,'M'),(5,'M'),(27,'M'),
      (41,'M')]),

    ('共情内核/包容力', '人际共情',
     [(37,'H'),(44,'H'),(31,'H'),(39,'H'),(48,'H'),(12,'H'),
      (29,'M'),(30,'M'),(13,'M'),(45,'M'),(41,'M'),
      (2,'L'),(0,'L'),(4,'L'),(43,'L')]),

    ('关系构建/维护', '人际共情',
     [(29,'H'),(30,'H'),(32,'H'),(31,'H'),(37,'H'),
      (13,'M'),(39,'M'),(27,'M'),(24,'M'),
      (2,'L'),(0,'L'),(4,'L'),(43,'L')]),

    ('协作沟通/社交智慧', '人际共情',
     [(13,'H'),(37,'H'),(32,'H'),(18,'H'),(45,'H'),
      (1,'M'),(29,'M'),(39,'M'),(42,'M'),
      (43,'L'),(4,'L')]),

    ('赋能型领导力', '人际共情',
     [(45,'H'),(46,'H'),(47,'H'),(48,'H'),(49,'H'),(13,'H'),
      (37,'M'),(12,'M'),(29,'M'),(39,'M'),
      (43,'L'),(41,'L')]),

    ('内在信念/品格', '意志行动',
     [(49,'H'),(45,'H'),(47,'H'),(46,'H'),(29,'H'),(15,'H'),
      (12,'M'),(13,'M'),(18,'M'),(19,'M'),(2,'M'),
      (43,'L'),(44,'L'),(41,'L')]),

    ('目标推进/执行力', '意志行动',
     [(13,'H'),(22,'H'),(29,'H'),(45,'H'),(46,'H'),(41,'H'),
      (12,'M'),(37,'M'),(39,'M'),(3,'M'),(1,'M'),
      (43,'L')]),

    ('自我管控/风险平衡', '意志行动',
     [(49,'H'),(45,'H'),(19,'H'),(33,'H'),(34,'H'),(12,'H'),
      (13,'M'),(29,'M'),(18,'M'),(47,'M'),(48,'M'),
      (43,'L'),(41,'L'),(44,'L')]),

    ('决策引领/变革', '意志行动',
     [(49,'H'),(45,'H'),(47,'H'),(46,'H'),(48,'H'),(12,'H'),
      (13,'M'),(17,'M'),(14,'M'),(32,'M'),(29,'M'),
      (43,'L'),(41,'L'),(44,'L')]),

    ('价值交付/服务', '领导职业',
     [(31,'H'),(44,'H'),(30,'H'),(39,'H'),(18,'H'),(42,'H'),
      (12,'M'),(29,'M'),(37,'M'),(41,'M'),(33,'M'),
      (4,'L'),(2,'L'),(43,'L')]),

    ('前瞻引领/领导力', '领导职业',
     [(49,'H'),(45,'H'),(46,'H'),(47,'H'),(48,'H'),(17,'H'),
      (12,'M'),(14,'M'),(27,'M'),
      (43,'L'),(41,'L'),(44,'L')]),

    ('落地统筹/执行', '领导职业',
     [(13,'H'),(45,'H'),(22,'H'),(41,'H'),(37,'H'),(46,'H'),
      (12,'M'),(29,'M'),(39,'M'),(42,'M'),
      (43,'L'),(44,'L')]),

    ('魅力调节因子', '领导职业',
     [(49,'H'),(45,'H'),(46,'H'),(29,'H'),(32,'H'),(48,'H'),
      (12,'M'),(13,'M'),(37,'M'),(39,'M'),(15,'M'),
      (4,'L'),(2,'L'),(43,'L')]),

    ('自我超越意愿', '超越动力',
     [(2,'H'),(15,'H'),(49,'H'),(12,'H'),(21,'H'),(15,'H'),
      (1,'M'),(6,'M'),(17,'M'),(19,'M'),(45,'M'),
      (43,'L'),(44,'L'),(41,'L')]),

    ('人生意义建构', '超越动力',
     [(49,'H'),(15,'H'),(14,'H'),(21,'H'),(45,'H'),
      (12,'M'),(39,'M'),(40,'M'),(18,'M'),
      (43,'L'),(44,'L'),(41,'L')]),

    ('利他/公共贡献', '超越动力',
     [(48,'H'),(39,'H'),(40,'H'),(49,'H'),(15,'H'),
      (12,'M'),(18,'M'),(27,'M'),
      (43,'L'),(4,'L'),(2,'L')]),

    ('超越性信念/灵性', '超越动力',
     [(49,'H'),(17,'H'),(19,'H'),(21,'H'),(14,'H'),
      (12,'M'),(45,'M'),(46,'M'),(15,'M'),
      (43,'L'),(44,'L'),(41,'L')]),

    ('欣赏美与卓越', '超越动力',
     [(6,'H'),(7,'H'),(8,'H'),(11,'H'),(10,'H'),(8,'H'),
      (12,'M'),(15,'M'),(27,'M'),(1,'M'),
      (43,'L'),(4,'L'),(2,'L'),(34,'L')]),

    ('超越性感恩/宽恕', '超越动力',
     [(49,'H'),(45,'H'),(48,'H'),(29,'H'),(31,'H'),(39,'H'),
      (12,'M'),(13,'M'),(37,'M'),(18,'M'),(41,'M'),
      (4,'L'),(2,'L'),(43,'L')]),
]

# 填充数据
for row_idx, (dim_name, category, relations) in enumerate(dimensions_data, 3):
    # 子维度名
    c = ws_summary.cell(row=row_idx, column=1, value=dim_name)
    c.font = Font(name='Arial', bold=True, size=9)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = border

    # 大类
    c = ws_summary.cell(row=row_idx, column=2, value=category)
    c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

    # 构建映射
    rel_map = {}
    for job_idx, level in relations:
        rel_map[job_idx] = level

    # 填充岗位列
    for job_idx in range(50):
        level = rel_map.get(job_idx, 'N')
        c = ws_summary.cell(row=row_idx, column=job_idx + 3, value=level)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        if level == 'H':
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        elif level == 'M':
            c.fill = PatternFill('solid', fgColor='FFEB9C')
        elif level == 'L':
            c.fill = PatternFill('solid', fgColor='FFCCCC')
        else:
            c.fill = PatternFill('solid', fgColor='F2F2F2')
            c.value = '—'

    ws_summary.row_dimensions[row_idx].height = 28

# 图例
legend_row = len(dimensions_data) + 5
ws_summary.cell(row=legend_row, column=1, value='图例').font = Font(name='Arial', bold=True, size=10)
legends = [
    ('H', 'C6EFCE', '核心匹配：成长型模式高分强烈推荐，等级型模式高分强烈不建议'),
    ('M', 'FFEB9C', '延伸匹配：该维度有正向加分，但非决定性因素'),
    ('L', 'FFCCCC', '不推荐：等级型模式高分适配，成长型模式高分不太适合'),
    ('—', 'F2F2F2', '不相关：该维度高低对该工作影响较小'),
]
for i, (sym, color, desc) in enumerate(legends):
    r = legend_row + 1 + i
    ws_summary.cell(row=r, column=1, value=sym).font = Font(name='Arial', size=10, bold=True)
    ws_summary.cell(row=r, column=1).fill = PatternFill('solid', fgColor=color)
    ws_summary.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=r, column=2, value=desc).font = Font(name='Arial', size=10)
    ws_summary.merge_cells(f'B{r}:G{r}')

ws_summary.freeze_panes = 'C3'

# ============ Sheet 3: 每个子维度详细表 ============
ws_detail = wb.create_sheet("子维度详细适配表")
ws_detail.sheet_view.showGridLines = False

ws_detail['A1'] = '27个子维度 — 详细适配分析与岗位推荐'
ws_detail['A1'].font = Font(name='Arial', bold=True, size=13)
ws_detail.merge_cells('A1:E1')
ws_detail['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_detail.row_dimensions[1].height = 28

# 表头
detail_headers = ['子维度', '所属大类', '成长型高分 → 推荐岗位', '等级型高分 → 推荐岗位', '核心说明']
for col, h in enumerate(detail_headers, 1):
    c = ws_detail.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws_detail.row_dimensions[2].height = 30

col_widths_detail = [20, 14, 40, 40, 40]
for i, w in enumerate(col_widths_detail, 1):
    ws_detail.column_dimensions[get_column_letter(i)].width = w

# 子维度详细数据
detail_data = [
    ('注意力/专注力', '认知能力',
     '算法工程师、硬件工程师、测试工程师、财务分析师、审计专员、数据录入员（需长时间专注）',
     '前端开发、UI设计、产品经理、电商运营、HRBP、客服、销售经理（需灵活切换多任务）',
     '高专注力→需深度思考的技术岗；低专注力（灵活型）→需多任务响应的运营/销售岗'),

    ('思维与洞察力', '认知能力',
     '产品经理、创新研究员、管理咨询顾问、商业分析师、行业研究员（洞察本质）',
     '数据录入员、行政、基础客服（按规则执行，不需深度洞察）',
     '高洞察力→分析决策类岗位；低洞察力→执行规则类岗位'),

    ('好奇心/学习力', '认知能力',
     '算法工程师、创新研究员、内容创作者、行业研究员、投资分析师、产品经理（持续学习）',
     '流水线操作、行政、数据录入员（任务重复，无需持续学习）',
     '强学习力→迭代快的新兴行业；弱学习力→稳定的传统行业'),

    ('信息分析/战略思维', '认知能力',
     '战略规划师、投资分析师、行业研究员、商业分析师、数字营销专员、财务分析师',
     '行政、基础客服、流水线操作（不需分析决策）',
     '强战略思维→分析决策层；弱战略思维→执行操作层'),

    ('元认知能力', '认知能力',
     '产品经理、管理咨询顾问、项目经理、运营总监、CEO创业者（自我迭代）',
     '流水线、数据录入、行政（固定流程，无需反思）',
     '强元认知→需持续迭代的复杂岗位；弱元认知→稳定规则岗位'),

    ('知识体系/跨界迁移', '认知能力',
     '创新研究员、产品经理、行业研究员、培训发展、CEO创业者（跨领域整合）',
     '流水线、数据录入、行政（单一技能即可）',
     '强跨界能力→创新/创业/研究；弱跨界→执行类岗位'),

    ('开放判断力', '认知能力',
     '管理咨询顾问、HRBP、OD组织发展、产品经理、HRD/CHO（多元协调）',
     '流水线操作、行政、基础客服（规则明确，无需协调冲突）',
     '强开放判断→冲突调解/咨询/HRBP；弱开放判断→合规执行岗'),

    ('创造力与创新', '认知能力',
     '产品经理、创新研究员、UI/UX设计师、品牌设计师、插画师、内容创作者',
     '审计、财务分析、行政、流水线操作（标准化为主）',
     '强创造力→创意/产品/内容岗；弱创造力→合规/审计/财务岗'),

    ('实践落地/反馈迭代', '认知能力',
     '项目经理、电商运营、销售经理、运营总监、产品经理（快速迭代拿结果）',
     '行政（流程固定，不需要频繁迭代）',
     '强落地迭代→业务冲锋型岗位；弱落地迭代→支持类岗位'),

    ('共情内核/包容力', '人际共情',
     'HRBP、客服、客户成功经理、培训发展、HRD/CHO、产品经理（理解他人）',
     '算法工程师、硬件工程师、数据录入员、独立研发岗（独立工作为主）',
     '强共情→服务/人资/产品岗；弱共情→技术深度岗'),

    ('关系构建/维护', '人际共情',
     '销售经理、大客户经理、BD商务拓展、客户成功经理、HRBP（拓展维护关系）',
     '算法工程师、硬件工程师、数据录入员（独立工作为主）',
     '强关系构建→对外拓展型岗位；弱关系构建→后台技术岗'),

    ('协作沟通/社交智慧', '人际共情',
     '项目经理、HRBP、BD商务拓展、管理咨询顾问、运营总监（协调多方）',
     '数据录入员、硬件工程师（独立工作为主）',
     '强协作沟通→协调型管理岗；弱协作沟通→技术执行岗'),

    ('赋能型领导力', '人际共情',
     '运营总监、市场总监、技术VP/CTO、HRD/CHO、CEO创业者、项目经理',
     '流水线、数据录入员、行政（独立执行，无需管理他人）',
     '强赋能→管理岗；弱赋能→执行层'),

    ('内在信念/品格', '意志行动',
     'CEO创业者、运营总监、技术VP、市场总监、销售经理、内容创作者（内核稳定）',
     '流水线操作、数据录入、行政、基础客服（稳定执行，无需自我驱动）',
     '强信念→创业/高管/销售；弱信念→稳定执行层'),

    ('目标推进/执行力', '意志行动',
     '项目经理、电商运营、销售经理、运营总监、市场总监、行政专员（高效执行）',
     '（较少天然适配，通常高执行力都较通用）',
     '强执行力→业务冲锋型岗位；通用适配，不存在天然不推荐'),

    ('自我管控/风险平衡', '意志行动',
     'CEO创业者、运营总监、投资分析师、财务分析师、审计、产品经理（风险管控）',
     '流水线、数据录入、行政、基础客服（低风险，无需管控复杂风险）',
     '强管控→金融/高管/战略岗；弱管控→低风险执行岗'),

    ('决策引领/变革', '意志行动',
     'CEO创业者、运营总监、技术VP、市场总监、HRD/CHO、产品经理（引领变革）',
     '流水线、数据录入、行政、基础客服（稳定为主，无需引领变革）',
     '强决策引领→高管/创业；弱决策引领→稳定执行层'),

    ('价值交付/服务', '领导职业',
     '客户成功经理、客服、大客户经理、培训发展、咨询顾问、法务专员（服务导向）',
     '算法工程师、硬件工程师、数据录入员（后台技术为主）',
     '强服务交付→客户服务岗；弱服务交付→后台技术岗'),

    ('前瞻引领/领导力', '领导职业',
     'CEO创业者、运营总监、市场总监、技术VP/CTO、HRD/CHO、战略规划师',
     '流水线、数据录入、行政、基础客服（执行层，不需引领方向）',
     '强领导力→高管/创始人；弱领导力→执行层'),

    ('落地统筹/执行', '领导职业',
     '项目经理、运营总监、电商运营、行政专员、HRBP、市场总监（统筹落地）',
     '流水线、数据录入、基础客服（单点执行，不需统筹）',
     '强统筹落地→项目管理/运营；弱统筹→单点执行岗'),

    ('魅力调节因子', '领导职业',
     'CEO创业者、运营总监、市场总监、销售经理、BD商务拓展、HRD/CHO（放大效果）',
     '算法工程师、硬件工程师、数据录入员（独立工作，影响较小）',
     '强魅力因子→管理/销售/BD；弱魅力→技术岗（不影响专业能力发挥）'),

    ('自我超越意愿', '超越动力',
     '算法工程师、内容创作者、创新研究员、CEO创业者、产品经理、行业研究员（内驱成长）',
     '流水线、数据录入、行政、基础客服（稳定即可，不需要主动成长）',
     '强自我超越→技术/创业/研究；弱自我超越→稳定执行层'),

    ('人生意义建构', '超越动力',
     'CEO创业者、内容创作者、创新研究员、行业研究员、运营总监（赋予工作意义）',
     '流水线、数据录入、基础客服（任务导向，不需要赋予深层意义）',
     '强意义建构→创业/内容/研究；弱意义建构→任务执行层'),

    ('利他/公共贡献', '超越动力',
     'HRD/CHO、培训发展、OD组织发展、CEO创业者、内容创作者（公共价值导向）',
     '流水线、数据录入、硬件工程师、算法工程师（个人专业为主，影响较小）',
     '强利他意愿→公益/HR/培训；弱利他→个人技术岗'),

    ('超越性信念/灵性', '超越动力',
     'CEO创业者、战略规划师、投资分析师、行业研究员、创新研究员（长期信念驱动）',
     '流水线、数据录入、行政、基础客服（短期任务导向，不需要信念驱动）',
     '强信念感→投资/战略/创业；弱信念→执行层'),

    ('欣赏美与卓越', '超越动力',
     'UI设计师、UX设计师、品牌设计师、插画师、视频剪辑师、工业设计师（审美驱动）',
     '流水线、数据录入、算法工程师、审计专员（不涉及审美，影响较小）',
     '强审美→设计/创意岗；弱审美→非设计/技术/财务岗'),

    ('超越性感恩/宽恕', '超越动力',
     'CEO创业者、运营总监、HRD/CHO、销售经理、客户成功经理、培训发展（情绪稳定）',
     '流水线、数据录入、硬件工程师、算法工程师（个人独立，影响较小）',
     '强感恩宽恕→高管/销售/HR；弱感恩宽恕→个人技术岗'),
]

for row_idx, (dim, cat, grow, level, note) in enumerate(detail_data, 3):
    for col_idx, val in enumerate([dim, cat, grow, level, note], 1):
        c = ws_detail.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = border
    ws_detail.row_dimensions[row_idx].height = 60

ws_detail.freeze_panes = 'A3'

wb.save(r'C:\Users\25890\Desktop\子维度工作关系表_个人成长名片.xlsx')
print('子维度工作关系表_个人成长名片.xlsx 创建成功！')
