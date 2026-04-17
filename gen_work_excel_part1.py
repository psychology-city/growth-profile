# -*- coding: utf-8 -*-
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

border = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))

with open(r'C:\Users\25890\.qclaw\workspace\growth-profile\work_data.json', encoding='utf-8') as f:
    jobs = json.load(f)

dims = [
    '注意力', '思维与洞察力', '好奇心与学习力', '信息分析与战略思维', '元认知能力',
    '知识体系与跨界迁移', '开放判断力', '创造力与创新', '实践落地与反馈迭代',
    '共情内核与包容力', '关系构建与维护', '协作沟通与社交智慧', '赋能型领导力',
    '内在信念与品格', '目标推进与执行力', '自我管控与风险平衡', '决策引领与变革',
    '价值交付与客户服务', '前瞻引领与战略思维', '落地统筹与项目管理', '魅力调节因子',
    '自我超越意愿', '人生意义建构', '利他与公共贡献', '超越性信念',
    '欣赏美与卓越', '超越性感恩与宽恕']

short_map = {
    '注意力': '注意力', '思维与洞察力': '思维洞察', '好奇心与学习力': '好奇心学习',
    '信息分析与战略思维': '信息分析', '元认知能力': '元认知',
    '知识体系与跨界迁移': '跨界迁移', '开放判断力': '开放判断',
    '创造力与创新': '创造力', '实践落地与反馈迭代': '实践落地',
    '共情内核与包容力': '共情包容', '关系构建与维护': '关系构建',
    '协作沟通与社交智慧': '协作沟通', '赋能型领导力': '赋能领导',
    '内在信念与品格': '内在信念', '目标推进与执行力': '目标推进',
    '自我管控与风险平衡': '自我管控', '决策引领与变革': '决策引领',
    '价值交付与客户服务': '价值交付', '前瞻引领与战略思维': '前瞻引领',
    '落地统筹与项目管理': '落地统筹', '魅力调节因子': '魅力因子',
    '自我超越意愿': '自我超越', '人生意义建构': '意义建构',
    '利他与公共贡献': '利他贡献', '超越性信念': '超越信念',
    '欣赏美与卓越': '欣赏美', '超越性感恩与宽恕': '超越感恩'}

cat_map = {
    '技术研发': '技术研发', '设计创意': '设计创意', '产品创新': '产品创新',
    '战略咨询': '战略咨询', '运营执行': '运营执行', '市场销售': '市场销售',
    '财务金融': '财务金融', '人力资源': '人力资源',
    '行政支持': '行政支持', '高级管理': '高级管理'}

cat_bg = {
    '技术研发': 'DAE8FC', '设计创意': 'D5E8D4', '产品创新': 'F8CECC',
    '战略咨询': 'FFF2CC', '运营执行': 'FFF8DC', '市场销售': 'E1D5E7',
    '财务金融': 'FCE4D6', '人力资源': 'D5E8D4',
    '行政支持': 'F5F5F5', '高级管理': 'F8CCE8'}

# Build dim -> jobs
dim_jobs = {d: [] for d in dims}
for j in jobs:
    for d in j['b']:
        if d in dim_jobs:
            dim_jobs[d].append(j['n'])

wb = Workbook()

# ===== Sheet 1: 总表 =====
ws1 = wb.active
ws1.title = '工作提升维度总表'
ws1.sheet_view.showGridLines = False

ws1['A1'] = '工作岗位能够提升的子维度总览'
ws1['A1'].font = Font(name='Arial', bold=True, size=14)
ws1.merge_cells('A1:G1')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

headers = ['序号', '工作岗位', '大类', '核心工作内容', '能提升的子维度（核心）', '提升等级', '提升方式说明']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 36
ws1.column_dimensions['E'].width = 38
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 42

for row_idx, j in enumerate(jobs, 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'
    data = [row_idx - 2, j['n'], j['c'], j['d'], '、'.join(j['b']), '核心提升', j['m']]
    for col_idx, val in enumerate(data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if col_idx == 1:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 2:
            c.font = Font(name='Arial', size=9, bold=True)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col_idx == 3:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=cat_bg.get(j['c'], bg))
        elif col_idx == 5:
            c.font = Font(name='Arial', size=9, color='1F497D')
            c.fill = PatternFill('solid', fgColor='EBF3FB')
        elif col_idx == 6:
            c.font = Font(name='Arial', size=9, bold=True, color='1F7A1F')
            c.fill = PatternFill('solid', fgColor='C6EFCE')
            c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
    ws1.row_dimensions[row_idx].height = 50

ws1.freeze_panes = 'A3'
print('Sheet1 done')

# ===== Sheet 2: 子维度 -> 工作 =====
ws2 = wb.create_sheet('子维度提升工作映射')
ws2.sheet_view.showGridLines = False

ws2['A1'] = '27个子维度 - 能够提升该维度的工作岗位推荐'
ws2['A1'].font = Font(name='Arial', bold=True, size=13)
ws2.merge_cells('A1:D1')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

for col, text in enumerate(['子维度', '大类', '能提升该维度的工作岗位', '典型提升方式'], 1):
    c = ws2.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 36
ws2.column_dimensions['D'].width = 38

for row_idx, dim in enumerate(dims, 3):
    cat = cat_map.get(dim, '其他')
    bg = cat_bg.get(cat, 'FFFFFF')
    jlist = dim_jobs.get(dim, [])

    c = ws2.cell(row=row_idx, column=1, value=dim)
    c.font = Font(name='Arial', size=9, bold=True)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = border

    c = ws2.cell(row=row_idx, column=2, value=cat)
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='top')
    c.border = border

    jlist_str = '、'.join(jlist) if jlist else '暂无数据'
    c = ws2.cell(row=row_idx, column=3, value=jlist_str)
    c.font = Font(name='Arial', size=9, color='1F497D')
    c.fill = PatternFill('solid', fgColor='EBF3FB')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    methods = [j['m'][:38] for j in jobs if dim in j['b']][:3]
    m_str = ' | '.join(methods) if methods else '暂无数据'
    c = ws2.cell(row=row_idx, column=4, value=m_str)
    c.font = Font(name='Arial', size=8)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border
    ws2.row_dimensions[row_idx].height = 60

ws2.freeze_panes = 'A3'
print('Sheet2 done')

# ===== Sheet 3: 矩阵 =====
ws3 = wb.create_sheet('提升关系矩阵')
ws3.sheet_view.showGridLines = False

ws3['A1'] = '50个工作岗位 x 27个子维度 - 提升关系矩阵'
ws3['A1'].font = Font(name='Arial', bold=True, size=12)
ws3.merge_cells('A1:AJ1')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 12
for i in range(3, 30):
    ws3.column_dimensions[get_column_letter(i)].width = 8

for col, text in enumerate(['工作岗位', '大类'], 1):
    c = ws3.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

for col, dim in enumerate(dims, 3):
    c = ws3.cell(row=2, column=col, value=short_map.get(dim, dim))
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws3.row_dimensions[2].height = 45

for row_idx, j in enumerate(jobs, 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'
    c = ws3.cell(row=row_idx, column=1, value=j['n'])
    c.font = Font(name='Arial', size=8, bold=True)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = border

    c = ws3.cell(row=row_idx, column=2, value=j['c'])
    c.font = Font(name='Arial', size=8)
    c.fill = PatternFill('solid', fgColor=cat_bg.get(j['c'], 'FFFFFF'))
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

    for col_idx, dim in enumerate(dims, 3):
        if dim in j['b']:
            c = ws3.cell(row=row_idx, column=col_idx, value='★')
            c.font = Font(name='Arial', size=9, color='1F7A1F', bold=True)
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        else:
            c = ws3.cell(row=row_idx, column=col_idx, value='')
            c.fill = PatternFill('solid', fgColor='F5F5F5')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws3.row_dimensions[row_idx].height = 22

lr = len(jobs) + 5
ws3.cell(row=lr, column=1, value='图例').font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=lr+1, column=1, value='★').font = Font(name='Arial', size=12, color='1F7A1F', bold=True)
ws3.cell(row=lr+1, column=1).fill = PatternFill('solid', fgColor='C6EFCE')
ws3.cell(row=lr+1, column=1).alignment = Alignment(horizontal='center')
ws3.cell(row=lr+1, column=2, value='该工作可显著提升此维度').font = Font(name='Arial', size=9)
ws3.merge_cells(f'B{lr+1}:G{lr+1}')
ws3.cell(row=lr+2, column=2, value='空白=该工作与该维度关联较弱，不作为主要提升路径').font = Font(name='Arial', size=9)
ws3.merge_cells(f'B{lr+2}:G{lr+2}')
ws3.freeze_panes = 'C3'
print('Sheet3 done')

# ===== 保存 =====
output = r'C:\Users\25890\Desktop\工作提升维度.xlsx'
wb.save(output)
print(f'Done! Saved to {output}')
print(f'Total jobs: {len(jobs)}')
for d in dims[:5]:
    print(f'  {d}: {len(dim_jobs[d])} jobs -> {dim_jobs[d]}')
