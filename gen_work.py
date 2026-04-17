# -*- coding: utf-8 -*-
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

border = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))

# 27个维度
dims_full = [
    '注意力', '思维与洞察力', '好奇心与学习力', '信息分析与战略思维',
    '元认知能力', '知识体系与跨界迁移', '开放判断力', '创造力与创新',
    '实践落地与反馈迭代', '共情内核与包容力', '关系构建与维护',
    '协作沟通与社交智慧', '赋能型领导力', '内在信念与品格',
    '目标推进与执行力', '自我管控与风险平衡', '决策引领与变革',
    '价值交付型服务能力', '前瞻引领型领导力', '落地统筹型执行能力',
    '魅力调节因子', '自我超越意愿', '人生意义建构',
    '利他与公共贡献', '超越性信念', '欣赏美与卓越', '超越性感恩与宽恕',
]

# 维度简称
dim_short = {
    '注意力': '注意力', '思维与洞察力': '思维洞察', '好奇心与学习力': '好奇心学习',
    '信息分析与战略思维': '信息分析', '元认知能力': '元认知', '知识体系与跨界迁移': '跨界迁移',
    '开放判断力': '开放判断', '创造力与创新': '创造力', '实践落地与反馈迭代': '实践落地',
    '共情内核与包容力': '共情包容', '关系构建与维护': '关系构建', '协作沟通与社交智慧': '协作沟通',
    '赋能型领导力': '赋能领导', '内在信念与品格': '内在信念', '目标推进与执行力': '目标推进',
    '自我管控与风险平衡': '自我管控', '决策引领与变革': '决策引领', '价值交付型服务能力': '价值交付',
    '前瞻引领型领导力': '前瞻领导', '落地统筹型执行能力': '落地统筹', '魅力调节因子': '魅力因子',
    '自我超越意愿': '自我超越', '人生意义建构': '意义建构', '利他与公共贡献': '利他贡献',
    '超越性信念': '超越信念', '欣赏美与卓越': '欣赏美', '超越性感恩与宽恕': '超越感恩',
}

# 维度类别
dim_cat = {
    '注意力': '认知', '思维与洞察力': '认知', '好奇心与学习力': '认知',
    '信息分析与战略思维': '认知', '元认知能力': '认知', '知识体系与跨界迁移': '认知',
    '开放判断力': '认知', '创造力与创新': '认知', '实践落地与反馈迭代': '认知',
    '共情内核与包容力': '人际', '关系构建与维护': '人际', '协作沟通与社交智慧': '人际',
    '赋能型领导力': '人际', '内在信念与品格': '意志', '目标推进与执行力': '意志',
    '自我管控与风险平衡': '意志', '决策引领与变革': '意志',
    '价值交付与客户服务': '领导', '前瞻引领型领导力': '领导', '落地统筹型执行能力': '领导',
    '魅力调节因子': '领导', '自我超越意愿': '超越', '人生意义建构': '超越',
    '利他与公共贡献': '超越', '超越性信念': '超越', '欣赏美与卓越': '超越',
    '超越性感恩与宽恕': '超越',
}

cat_colors = {
    '认知': 'DAE8FC', '人际': 'D5E8D4', '意志': 'FFF2CC',
    '领导': 'F8CECC', '超越': 'E1D5E7', '其他': 'F5F5F5',
}

dim_defs = {
    '注意力': '指长时间保持专注、不被干扰分散的能力，是深度工作的核心前提',
    '思维与洞察力': '透过现象看本质，快速抓住核心规律的能力，是问题分析和决策判断的基础',
    '好奇心与学习力': '主动探索未知、持续学习的内驱力，是知识积累和能力迭代的核心燃料',
    '信息分析与战略思维': '对复杂信息进行结构化分析、形成全局判断的能力，是战略决策的认知基础',
    '元认知能力': '对自己思维过程的监控和调节能力，是自我迭代和高效学习的元能力',
    '知识体系与跨界迁移': '构建系统知识框架、将一个领域的知识迁移到其他领域的能力',
    '开放判断力': '接纳不同观点、容纳矛盾信息、不被偏见局限的能力，是多元思维的基础',
    '创造力与创新': '产生新颖且有价值想法的能力，是产品创新和内容创作的核心驱动',
    '实践落地与反馈迭代': '将想法转化为行动、基于反馈快速优化的能力，是从知道到做到的关键',
    '共情内核与包容力': '感知和理解他人情绪与立场的能力，是人际沟通和客户服务的情感基础',
    '关系构建与维护': '主动拓展和维护有价值关系的能力，是资源整合和协作网络的核心',
    '协作沟通与社交智慧': '在多人协作中有效表达，协调分歧、达成共识的能力',
    '赋能型领导力': '通过支持他人成长来实现团队目标的能力，是从个人贡献者到管理者的关键跃迁',
    '内在信念与品格': '在压力和诱惑下坚守核心价值观的能力，是长期信任和人品的基石',
    '目标推进与执行力': '将目标转化为持续行动、克服阻力拿到结果的能力',
    '自我管控与风险平衡': '控制冲动，管理情绪、平衡风险的能力，是成熟决策的行为保障',
    '决策引领与变革': '在不确定性中做出果断决策并推动落地的能力，是管理者的核心职责',
    '价值交付与客户服务': '以客户需求为中心、交付超出预期价值的意识和行为模式',
    '前瞻引领与战略思维': '为团队指明方向、凝聚人心、引领变革的领导特质',
    '落地统筹与项目管理': '将模糊目标拆解为可执行动作，协调多方资源、确保落地的综合能力',
    '魅力调节因子': '通过真诚、幽默、感恩等社交魅力放大其他能力落地效果的能力',
    '自我超越意愿': '不满足于现状、主动突破舒适区、持续追求成长的内驱力',
    '人生意义建构': '赋予工作和生活超越功利目标的意义感，是持久动力的深层来源',
    '利他与公共贡献': '愿意为他人、行业、社会创造价值的意识，是领导力和影响力的社会基础',
    '超越性信念': '相信宏大规律和长期价值、愿意承受短期不确定性的信念感',
    '欣赏美与卓越': '感知美、追求卓越、欣赏高品质事物的意识，是审美和创意的基础',
    '超越性感恩与宽恕': '对拥有的一切心怀感恩、对过去释怀的心态，是情绪稳定的深层来源',
}

# 50个工作岗位数据：(名称, 大类, 核心工作内容, 能提升该维度的子维度列表, 提升等级, 提升方式说明, 维度说明)
# 工作提升的子维度（从之前生成的job文件中提取的核心数据）
jobs = [
    # 技术研发类
    ('后端开发工程师', '技术研发类', '服务器端系统设计与开发、数据库架构、API接口、性能优化',
     ['注意力', '思维与洞察力', '目标推进与执行力', '实践落地与反馈迭代', '元认知能力'],
     '核心提升', '代码编写需要持续专注；系统设计培养思维洞察；项目开发培养目标推进和落地能力'),
    ('前端开发工程师', '技术研发类', 'Web界面开发、框架搭建、页面性能优化、交互动效实现',
     ['注意力', '审美与卓越', '好奇心与学习力', '创造力与创新', '协作沟通'],
     '核心提升', '页面开发需要专注力和审美感知；框架学习培养好奇心；团队协作培养沟通能力'),
    ('算法工程师', '技术研发类', '机器学习模型研发、算法优化与工程化',
     ['注意力', '好奇心与学习力', '思维与洞察力', '信息分析与战略思维'],
     '核心提升', '模型研发需要深度专注和学习力；算法优化培养洞察力和分析能力'),
    ('数据分析师', '技术研发类', '数据清洗处理、指标体系搭建、可视化报告输出',
     ['信息分析与战略思维', '思维与洞察力', '好奇心与学习力', '协作沟通'],
     '核心提升', '数据分析培养洞察力和分析能力；报告输出培养沟通能力'),
    ('硬件工程师', '技术研发类', '电路设计、芯片研发、硬件测试与调试',
     ['注意力', '目标推进与执行力', '思维与洞察力'],
     '核心提升', '硬件调试需要持续专注；项目周期培养目标推进；问题排查培养洞察力'),
    ('测试工程师', '技术研发类', '测试计划制定、用例编写、功能性能测试、Bug跟踪',
     ['注意力', '目标推进与执行力', '协作沟通'],
     '核心提升', '测试执行需要持续专注；测试计划培养目标推进；跨团队沟通培养协作能力'),

    # 设计创意类
    ('UI设计师', '设计创意类', '产品界面视觉设计、图标绘制、设计规范制定',
     ['审美与卓越', '注意力', '创造力与创新', '协作沟通'],
     '核心提升', '界面设计培养审美感知；用户研究培养协作沟通；创新设计培养创造力'),
    ('UX设计师', '设计创意类', '用户研究、交互流程设计、原型制作与用户测试',
     ['共情内核与包容力', '好奇心与学习力', '思维与洞察力'],
     '核心提升', '用户研究直接培养共情能力；需求分析培养洞察力'),
    ('品牌设计师', '设计创意类', '品牌视觉体系搭建、宣传物料设计、广告创意视觉',
     ['审美与卓越', '创造力与创新', '目标推进与执行力', '协作沟通'],
     '核心提升', '品牌设计培养审美和创新；项目执行培养目标推进和协作'),
    ('工业设计师', '设计创意类', '产品外观设计、材料工艺选择、手板制作',
     ['审美与卓越', '好奇心与学习力', '目标推进与执行力', '实践落地与反馈迭代'],
     '核心提升', '工业设计培养审美和好奇心；手板制作培养落地能力'),
    ('视频剪辑师', '设计创意类', '视频剪辑、特效合成、调色、音频处理与包装',
     ['审美与卓越', '注意力', '目标推进与执行力'],
     '核心提升', '剪辑需要专注力和审美；项目交付培养目标推进能力'),
    ('插画师', '设计创意类', '商业插画绘制、IP形象设计、图文内容创作',
     ['审美与卓越', '创造力与创新', '好奇心与学习力'],
     '核心提升', '插画创作培养审美和创造力；商业项目培养对市场和用户的理解'),

    # 产品创新类
    ('产品经理', '产品创新类', '需求调研分析、产品设计、跨部门沟通、上线跟进与迭代优化',
     ['思维与洞察力', '目标推进与执行力', '协作沟通与社交智慧', '共情内核与包容力', '决策引领与变革'],
     '核心提升', '需求分析培养洞察力；项目管理培养目标推进；跨部门协作培养沟通；用户研究培养共情'),
    ('项目经理', '产品创新类', '项目全生命周期管理、里程碑制定、风险管控、干系人管理',
     ['目标推进与执行力', '协作沟通与社交智慧', '自我管控与风险平衡', '决策引领与变革'],
     '核心提升', '项目管理直接培养目标推进和落地统筹能力；风险管理培养自我管控'),
    ('创新研究员', '产品创新类', '行业前沿研究、创新趋势洞察、新技术评估、创新项目孵化',
     ['好奇心与学习力', '审美与卓越', '超越性信念'],
     '核心提升', '前沿研究培养好奇心和持续学习；长期追踪培养超越信念'),
    ('内容创作者', '产品创新类', '内容选题策划、文案撰写、视频拍摄剪辑、粉丝运营与互动',
     ['审美与卓越', '创造力与创新', '好奇心与学习力', '超越性感恩与宽恕'],
     '核心提升', '内容创作培养审美和创造力；持续输出培养好奇心和超越感恩'),
    ('知识产权专员', '产品创新类', '专利申请管理、知识产权维权、商标注册、技术合同审核',
     ['注意力', '开放判断力', '目标推进与执行力'],
     '核心提升', '专业审查培养注意力和判断力；流程管理培养目标推进能力'),

    # 战略咨询类
    ('战略规划师', '战略咨询类', '企业战略制定、行业趋势研究、竞争分析、战略落地跟踪',
     ['信息分析与战略思维', '思维与洞察力', '超越性信念', '协作沟通与社交智慧'],
     '核心提升', '战略分析培养洞察力和信息分析能力；战略落地培养协作沟通'),
    ('管理咨询顾问', '战略咨询类', '管理诊断、方案设计、实施辅导',
     ['思维与洞察力', '协作沟通与社交智慧', '好奇心与学习力', '开放判断力'],
     '核心提升', '管理咨询培养系统洞察力；客户沟通培养协作能力；跨行业研究培养开放判断力'),
    ('投资分析师', '战略咨询类', '行业研究、项目尽调、财务建模、投资报告撰写、投后管理',
     ['信息分析与战略思维', '好奇心与学习力', '自我管控与风险平衡'],
     '核心提升', '投资分析培养信息分析和洞察力；风控培养自我管控能力'),
    ('商业分析师', '战略咨询类', '业务数据分析、经营诊断、流程优化建议、方案推动落地',
     ['信息分析与战略思维', '目标推进与执行力', '协作沟通与社交智慧'],
     '核心提升', '数据分析培养洞察力；方案落地培养目标推进和协作能力'),
    ('行业研究员', '战略咨询类', '深度行业研究、研报撰写、路演与客户服务',
     ['好奇心与学习力', '信息分析与战略思维', '超越性信念'],
     '核心提升', '行业研究培养好奇心和信息分析；长期跟踪培养超越信念'),

    # 运营执行类
    ('电商运营', '运营执行类', '店铺日常运营、活动策划、流量获取转化、数据监控与复盘',
     ['目标推进与执行力', '协作沟通与社交智慧', '自我超越意愿'],
     '核心提升', '电商运营直接培养目标推进和执行力；活动策划培养协作沟通'),
    ('用户运营', '运营执行类', '用户分层管理、增长策略、活跃留存提升、会员体系搭建',
     ['共情内核与包容力', '好奇心与学习力', '目标推进与执行力'],
     '核心提升', '用户运营培养共情能力；增长目标培养目标推进力'),
    ('内容运营', '运营执行类', '内容选题生产、分发推广、数据分析优化、创作者生态维护',
     ['审美与卓越', '好奇心与学习力', '目标推进与执行力'],
     '核心提升', '内容运营培养审美和目标推进；数据分析培养信息分析能力'),
    ('供应链专员', '运营执行类', '采购管理、库存控制、物流协调、供应商关系维护',
     ['目标推进与执行力', '自我管控与风险平衡', '协作沟通与社交智慧'],
     '核心提升', '供应链管理培养目标推进和风险管控；多方协调培养协作能力'),
    ('质量管理专员', '运营执行类', '质量标准制定与执行、检验检测、质量问题处理与改进',
     ['注意力', '目标推进与执行力', '开放判断力'],
     '核心提升', '质量检查培养注意力和判断力；体系建设培养目标推进能力'),

    # 市场销售类
    ('品牌营销经理', '市场销售类', '品牌战略制定与执行、市场推广计划、媒介合作、营销活动策划',
     ['审美与卓越', '协作沟通与社交智慧', '目标推进与执行力', '超越性信念'],
     '核心提升', '品牌管理培养审美和超越信念；活动策划培养目标推进能力'),
    ('数字营销专员', '市场销售类', 'SEM/SEO投放、社媒广告、数据分析、转化漏斗优化',
     ['信息分析与战略思维', '目标推进与执行力', '好奇心与学习力'],
     '核心提升', '数据分析培养洞察力；投放优化培养目标推进能力'),
    ('销售经理', '市场销售类', '客户开发维护、销售目标达成、商务谈判、合同签订与回款管理',
     ['共情内核与包容力', '目标推进与执行力', '超越性信念'],
     '核心提升', '客户关系培养共情能力；业绩压力培养目标推进和超越信念'),
    ('大客户经理', '市场销售类', '大客户开拓与深度维护、定制化解决方案、战略合作谈判',
     ['共情内核与包容力', '超越信念', '协作沟通与社交智慧', '目标推进与执行力'],
     '核心提升', '大客户管理培养共情和超越信念；解决方案培养系统思维和目标推进'),
    ('客户成功经理', '市场销售类', '客户onboarding、产品价值传递、健康度管理、续费与增购',
     ['共情内核与包容力', '超越性感恩与宽恕', '目标推进与执行力'],
     '核心提升', '客户服务直接培养共情能力和超越感恩；客户成功体系培养目标推进能力'),
    ('BD商务拓展', '市场销售类', '商务合作机会挖掘、合作方案设计、合同谈判、资源整合',
     ['协作沟通与社交智慧', '目标推进与执行力', '超越性信念'],
     '核心提升', '商务谈判培养协作沟通；合作落地培养目标推进能力；长期关系培养超越信念'),

    # 财务金融类
    ('财务分析师', '财务金融类', '财务预算编制、经营分析、报表解读与建议、财务模型搭建',
     ['信息分析与战略思维', '注意力', '目标推进与执行力'],
     '核心提升', '财务分析培养信息分析和洞察力；预算管理培养目标推进能力'),
    ('审计专员', '财务金融类', '财务审计、内部控制评估、合规检查、审计报告撰写',
     ['注意力', '开放判断力', '目标推进与执行力'],
     '核心提升', '审计工作培养注意力和开放判断力；审计流程培养目标推进能力'),
    ('投融资专员', '财务金融类', '融资对接、投资机构关系维护、商业计划书撰写、估值谈判',
     ['超越性信念', '好奇心与学习力', '协作沟通与社交智慧'],
     '核心提升', '投融资工作培养超越信念；商务沟通培养协作能力；行业研究培养好奇心'),
    ('税务专员', '财务金融类', '税务筹划、税务申报缴纳、税务稽查应对、优惠政策申请',
     ['注意力', '目标推进与执行力', '超越性信念'],
     '核心提升', '税务合规培养注意力和超越信念；申报流程培养目标推进能力'),

    # 人力资源类
    ('HRBP', '人力资源类', '人才战略制定、招聘配置、绩效管理、员工关系维护',
     ['共情内核与包容力', '协作沟通与社交智慧', '超越性感恩与宽恕', '目标推进与执行力'],
     '核心提升', 'HR工作直接培养共情和超越感恩；人才管理培养目标推进能力'),
    ('招聘专员', '人力资源类', '招聘需求理解、人才画像制定、渠道运营、面试组织与候选人跟进',
     ['共情内核与包容力', '好奇心与学习力', '目标推进与执行力'],
     '核心提升', '人才甄别培养共情能力；招聘流程培养目标推进能力'),
    ('培训发展专员', '人力资源类', '培训体系搭建、课程设计开发、培训项目运营、效果评估',
     ['超越性信念', '好奇心与学习力', '协作沟通与社交智慧'],
     '核心提升', '课程开发培养超越信念和好奇心；培训交付培养协作沟通能力'),
    ('OD组织发展专员', '人力资源类', '组织诊断、变革管理、文化落地、领导力发展项目设计',
     ['超越性信念', '好奇心与学习力', '协作沟通与社交智慧', '目标推进与执行力'],
     '核心提升', 'OD工作培养超越信念和目标推进；组织变革培养协作沟通能力'),

    # 行政支持类
    ('行政专员', '行政支持类', '日常行政事务处理、办公环境管理、活动组织、后勤保障',
     ['目标推进与执行力', '超越性感恩与宽恕', '协作沟通与社交智慧'],
     '核心提升', '行政事务培养目标推进能力；多任务协调培养协作沟通；服务意识培养超越感恩'),
    ('法务专员', '行政支持类', '合同审核、法律咨询、诉讼跟进、合规管理',
     ['注意力', '开放判断力', '目标推进与执行力'],
     '核心提升', '法务审查培养注意力和开放判断；合规管理培养目标推进能力'),
    ('数据录入员', '行政支持类', '数据采集录入、系统维护、数据核查清洗、报表汇总',
     ['注意力', '目标推进与执行力'],
     '核心提升', '数据录入培养注意力和目标推进；质量控制培养自我管控能力'),
    ('客服专员', '行政支持类', '客户咨询解答、投诉处理、问题反馈记录、满意度维护',
     ['共情内核与包容力', '超越性感恩与宽恕', '目标推进与执行力'],
     '核心提升', '客户服务直接培养共情和超越感恩；投诉处理培养目标推进能力'),

    # 高级管理类
    ('运营总监', '高级管理类', '运营战略制定与执行、多部门协调、核心KPI达成、组织效能提升',
     ['目标推进与执行力', '决策引领与变革', '超越性信念', '利他与公共贡献'],
     '核心提升', '全面运营管理培养目标推进和决策引领；团队管理培养超越信念和利他意愿'),
    ('市场总监', '高级管理类', '市场营销战略制定、品牌管理、团队建设、预算控制、市场份额增长',
     ['审美与卓越', '目标推进与执行力', '超越性信念', '协作沟通与社交智慧'],
     '核心提升', '品牌管理培养审美和超越信念；团队管理培养协作沟通和目标推进能力'),
    ('技术VP/CTO', '高级管理类', '技术战略制定、研发团队管理、技术架构决策、技术品牌建设',
     ['好奇心与学习力', '目标推进与执行力', '协作沟通与社交智慧', '审美与卓越'],
     '核心提升', '技术管理培养好奇心和目标推进；团队领导培养协作沟通和审美能力'),
    ('HRD/CHO', '高级管理类', '人力资源战略、OD与文化建设、高管人才管理、雇主品牌、HR数字化',
     ['超越性信念', '目标推进与执行力', '协作沟通与社交智慧', '利他与公共贡献'],
     '核心提升', '人才战略培养超越信念；文化建设培养利他意愿；体系搭建培养目标推进能力'),
    ('CEO创业者', '高级管理类', '公司战略文化制定、高管团队建设、重大决策拍板、资本运作与对外关系',
     ['目标推进与执行力', '超越性信念', '好奇心与学习力', '利他与公共贡献'],
     '核心提升', '创业历程培养全面能力发展；使命驱动培养超越信念和利他意愿'),
]

print(f'Jobs count: {len(jobs)}')

# Build dim -> jobs mapping
dim_jobs = {d: [] for d in dims_full}
for job in jobs:
    for dim in job[4]:  # job[4] = boost_dims
        if dim in dim_jobs:
            dim_jobs[dim].append(job[0])

# ============== Sheet 1: 工作提升维度总表 ==============
wb = Workbook()
ws1 = wb.active
ws1.title = '工作提升维度总表'
ws1.sheet_view.showGridLines = False
ws1['A1'] = '工作岗位能够提升的子维度总览'
ws1['A1'].font = Font(name='Arial', bold=True, size=13)
ws1.merge_cells('A1:G1')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 28

headers = ['序号', '工作岗位', '大类', '核心工作内容', '能提升的子维度（核心）', '提升等级', '提升方式说明']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

for col, w in enumerate([6, 18, 12, 38, 36, 10, 42], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

for row_idx, job in enumerate(jobs, 3):
    name, cat, content, boost_dims, level, method, _ = job
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'
    data = [row_idx-2, name, cat, content, '、'.join(boost_dims), level, method]
    for col_idx, val in enumerate(data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if col_idx == 1:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center')
        elif col_idx == 2:
            c.font = Font(name='Arial', size=9, bold=True)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col_idx == 3:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col_idx == 5:
            c.font = Font(name='Arial', size=9, color='1F497D')
            c.fill = PatternFill('solid', fgColor='EBF3FB')
        elif col_idx == 6:
            c.font = Font(name='Arial', size=9, bold=True, color='1F7A1F')
            c.fill = PatternFill('solid', fgColor='C6EFCE')
            c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
    ws1.row_dimensions[row_idx].height = 50

ws1.freeze_panes = 'A3'

# ============== Sheet 2: 子维度 -> 工作映射 ==============
ws2 = wb.create_sheet('子维度提升工作映射')
ws2.sheet_view.showGridLines = False
ws2['A1'] = '27个子维度 - 能够提升该维度的工作岗位推荐'
ws2['A1'].font = Font(name='Arial', bold=True, size=13)
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

for col, text in enumerate(['子维度', '大类', '维度说明', '能提升该维度的工作岗位'], 1):
    c = ws2.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 36
ws2.column_dimensions['D'].width = 36

for row_idx, dim in enumerate(dims_full, 3):
    cat = dim_cat.get(dim, '其他')
    bg = cat_colors.get(cat, 'FFFFFF')
    jlist = dim_jobs.get(dim, [])

    c = ws2.cell(row=row_idx, column=1, value=dim)
    c.font = Font(name='Arial', size=9, bold=True)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = border

    c = ws2.cell(row=row_idx, column=2, value=cat)
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='top')
    c.border = border

    c = ws2.cell(row=row_idx, column=3, value=dim_defs.get(dim, ''))
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    job_str = '、'.join(jlist) if jlist else '暂无数据'
    c = ws2.cell(row=row_idx, column=4, value=job_str)
    c.font = Font(name='Arial', size=9, color='1F497D')
    c.fill = PatternFill('solid', fgColor='EBF3FB')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    ws2.row_dimensions[row_idx].height = 70

ws2.freeze_panes = 'A3'

# ============== Sheet 3: 关系矩阵 ==============
ws3 = wb.create_sheet('提升关系矩阵')
ws3.sheet_view.showGridLines = False
ws3['A1'] = '50个工作岗位 x 27个子维度 - 提升关系矩阵'
ws3['A1'].font = Font(name='Arial', bold=True, size=12)
ws3.merge_cells('A1:AJ1')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 12
for i in range(3, 30):
    ws3.column_dimensions[get_column_letter(i)].width = 8

for col, text in enumerate(['工作岗位', '大类'], 1):
    c = ws3.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

for col, dim in enumerate(dims_full, 3):
    short = dim_short.get(dim, dim)
    c = ws3.cell(row=2, column=col, value=short)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws3.row_dimensions[2].height = 45

job_names = [j[0] for j in jobs]
job_cats = [j[1] for j in jobs]
job_boost = [j[4] for j in jobs]  # boost dims

for row_idx, (job_name, job_cat, job_content, boost_dims, level, method, dim_desc) in enumerate(jobs, 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'
    for col, text in [(1, job_name), (2, job_cat)]:
        c = ws3.cell(row=row_idx, column=col, value=text)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='center')
        if col == 1:
            c.font = Font(name='Arial', size=8, bold=True)
            c.fill = PatternFill('solid', fgColor=bg)
        else:
            c.font = Font(name='Arial', size=8)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center')
    for col_idx, dim in enumerate(dims_full, 3):
        if dim in boost_dims:
            c = ws3.cell(row=row_idx, column=col_idx, value='★')
            c.font = Font(name='Arial', size=9, color='1F7A1F', bold=True)
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        else:
            c = ws3.cell(row=row_idx, column=col_idx, value='')
            c.fill = PatternFill('solid', fgColor='F5F5F5')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws3.row_dimensions[row_idx].height = 22

# 图例
lr = len(jobs) + 5
ws3.cell(row=lr, column=1, value='图例').font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=lr+1, column=1, value='★').font = Font(name='Arial', size=12, color='1F7A1F', bold=True)
ws3.cell(row=lr+1, column=1).fill = PatternFill('solid', fgColor='C6EFCE')
ws3.cell(row=lr+1, column=1).alignment = Alignment(horizontal='center')
ws3.cell(row=lr+1, column=2, value='该工作可显著提升此维度').font = Font(name='Arial', size=9)
ws3.merge_cells(f'B{lr+1}:F{lr+1}')
ws3.cell(row=lr+2, column=2, value='空白=该工作与该维度关联较弱').font = Font(name='Arial', size=9)
ws3.merge_cells(f'B{lr+2}:F{lr+2}')

ws3.freeze_panes = 'C3'

# ============== 保存 ==============
output = r'C:\Users\25890\Desktop\工作提升维度.xlsx'
wb.save(output)
print(f'Saved: {output}')
