# -*- coding: utf-8 -*-
# Read detail_data.json and auto-generate the hobby boost Excel
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open(r'C:\Users\25890\.qclaw\workspace\growth-profile\detail_data.json', 'r', encoding='utf-8') as f:
    detail = json.load(f)

# Sub-dimension full names
dims_full = [
    '注意力', '思维与洞察力', '好奇心与学习力', '信息分析与战略思维',
    '元认知能力', '知识体系与跨界迁移', '开放判断力', '创造力与创新',
    '实践落地与反馈迭代', '共情内核与包容力', '关系构建与维护',
    '协作沟通与社交智慧', '赋能型领导力', '内在信念与品格',
    '目标推进与执行力', '自我管控与风险平衡', '决策引领与变革',
    '价值交付型服务能力', '前瞻引领型领导力', '落地统筹型执行能力',
    '魅力调节因子', '自我超越意愿', '人生意义建构',
    '利他与公共贡献', '超越性信念', '欣赏美与卓越', '超越性感恩与宽恕',
]

# Short name for display
short_map = {
    '注意力': '注意力', '思维与洞察力': '思维洞察', '好奇心与学习力': '好奇心学习',
    '信息分析与战略思维': '信息分析', '元认知能力': '元认知', '知识体系与跨界迁移': '跨界迁移',
    '开放判断力': '开放判断', '创造力与创新': '创造力', '实践落地与反馈迭代': '实践落地',
    '共情内核与包容力': '共情内核', '关系构建与维护': '关系构建', '协作沟通与社交智慧': '协作沟通',
    '赋能型领导力': '赋能领导', '内在信念与品格': '内在信念', '目标推进与执行力': '目标推进',
    '自我管控与风险平衡': '自我管控', '决策引领与变革': '决策引领', '价值交付型服务能力': '价值交付',
    '前瞻引领型领导力': '前瞻领导', '落地统筹型执行能力': '落地统筹', '魅力调节因子': '魅力调节',
    '自我超越意愿': '自我超越', '人生意义建构': '意义建构', '利他与公共贡献': '利他贡献',
    '超越性信念': '超越信念', '欣赏美与卓越': '欣赏美', '超越性感恩与宽恕': '超越感恩',
}

# Category
cat_map = {
    '注意力': '认知', '思维与洞察力': '认知', '好奇心与学习力': '认知',
    '信息分析与战略思维': '认知', '元认知能力': '认知', '知识体系与跨界迁移': '认知',
    '开放判断力': '认知', '创造力与创新': '认知', '实践落地与反馈迭代': '认知',
    '共情内核与包容力': '人际', '关系构建与维护': '人际', '协作沟通与社交智慧': '人际',
    '赋能型领导力': '人际', '内在信念与品格': '意志', '目标推进与执行力': '意志',
    '自我管控与风险平衡': '意志', '决策引领与变革': '意志', '价值交付型服务能力': '领导',
    '前瞻引领型领导力': '领导', '落地统筹型执行能力': '领导', '魅力调节因子': '领导',
    '自我超越意愿': '超越', '人生意义建构': '超越', '利他与公共贡献': '超越',
    '超越性信念': '超越', '欣赏美与卓越': '超越', '超越性感恩与宽恕': '超越',
}

cat_colors = {
    '认知': 'DAE8FC', '人际': 'D5E8D4', '意志': 'FFF2CC',
    '领导': 'F8CECC', '超越': 'E1D5E7',
}

# Parse the core dimensions from detail_data.json
# detail[i][2] = core dimensions (e.g. '注意力、内在信念、意志力、自我管控、超越感恩')
# detail[i][3] = extended dimensions
# detail[i][4] = growth suggestion
# detail[i][5] = level-type suggestion

def parse_dims(dim_string):
    """Parse comma-separated dimension names"""
    return [d.strip() for d in dim_string.split('、') if d.strip()]

# Build hobby list
hobbies = []
for item in detail:
    hobby = item[0]
    cat = item[1]
    core_dims = parse_dims(item[2])
    extend_dims = parse_dims(item[3])
    growth = item[4]
    level_sug = item[5]
    hobbies.append({
        'name': hobby,
        'cat': cat,
        'core_dims': core_dims,
        'extend_dims': extend_dims,
        'growth': growth,
        'level_sug': level_sug,
    })

# Build dim -> hobbies mapping
dim_hobbies = {d: [] for d in dims_full}
for h in hobbies:
    for d in h['core_dims']:
        if d in dim_hobbies:
            dim_hobbies[d].append(h['name'])

# Sub-dimension definitions
dim_defs = {
    '注意力': '指长时间保持专注、不被干扰分散的能力，是深度工作的核心前提',
    '思维与洞察力': '透过现象看本质、快速抓住核心规律的能力，是问题分析和决策判断的基础',
    '好奇心与学习力': '主动探索未知、持续学习的内驱力，是知识积累和能力迭代的核心燃料',
    '信息分析与战略思维': '对复杂信息进行结构化分析、形成全局判断的能力，是战略决策的认知基础',
    '元认知能力': '对自己思维过程的监控和调节能力，是自我迭代和高效学习的元能力',
    '知识体系与跨界迁移': '构建系统知识框架、将一个领域的知识迁移到其他领域的能力',
    '开放判断力': '接纳不同观点、容纳矛盾信息、不被偏见局限的能力，是多元思维的基础',
    '创造力与创新': '产生新颖且有价值想法的能力，是产品创新和内容创作的核心驱动',
    '实践落地与反馈迭代': '将想法转化为行动、基于反馈快速优化的能力，是从知道到做到的关键',
    '共情内核与包容力': '感知和理解他人情绪与立场的能力，是人际沟通和客户服务的情感基础',
    '关系构建与维护': '主动拓展和维护有价值关系的能力，是资源整合和协作网络的核心',
    '协作沟通与社交智慧': '在多人协作中有效表达，协调分歧、达成共识的能力',
    '赋能型领导力': '通过支持他人成长来实现团队目标的能力，是从个人贡献者到管理者的关键跃迁',
    '内在信念与品格': '在压力和诱惑下坚守核心价值观的能力，是长期信任和人品的基石',
    '目标推进与执行力': '将目标转化为持续行动、克服阻力拿到结果的能力',
    '自我管控与风险平衡': '控制冲动、管理情绪、平衡风险的能力，是成熟决策的行为保障',
    '决策引领与变革': '在不确定性中做出果断决策并推动落地的能力，是管理者的核心职责',
    '价值交付型服务能力': '以客户需求为中心、交付超出预期的价值的意识和行为模式',
    '前瞻引领型领导力': '为团队指明方向、凝聚人心、引领变革的领导特质',
    '落地统筹型执行能力': '将模糊目标拆解为可执行动作，协调多方资源、确保落地的综合能力',
    '魅力调节因子': '通过真诚、幽默、感恩等社交魅力放大其他能力落地效果的能力',
    '自我超越意愿': '不满足于现状、主动突破舒适区、持续追求成长的内驱力',
    '人生意义建构': '赋予工作和生活超越功利目标的意义感，是持久动力的深层来源',
    '利他与公共贡献': '愿意为他人、行业、社会创造价值的意识，是领导力和影响力的社会基础',
    '超越性信念': '相信宏大规律和长期价值、愿意承受短期不确定性的信念感',
    '欣赏美与卓越': '感知美、追求卓越、欣赏高品质事物的意识，是审美和创意的基础',
    '超越性感恩与宽恕': '对已拥有的一切心怀感恩、对过去释怀放下的心态，是情绪稳定的深层来源',
}

wb = Workbook()
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# =============== Sheet 1: 总表 ===============
ws1 = wb.active
ws1.title = '爱好提升维度总表'
ws1.sheet_view.showGridLines = False

ws1['A1'] = '兴趣爱好能够提升的子维度总览'
ws1['A1'].font = Font(name='Arial', bold=True, size=14)
ws1.merge_cells('A1:H1')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

headers1 = ['序号', '兴趣爱好', '大类', '能提升的子维度（核心）', '提升等级', '提升方式说明', '等级型高分者适配方向']
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

col_widths1 = [6, 14, 12, 38, 10, 44, 40]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for row_idx, h in enumerate(hobbies, 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'

    data = [row_idx-2, h['name'], h['cat'], '、'.join(h['core_dims']), '核心提升', h['growth'], h['level_sug']]
    for col_idx, val in enumerate(data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if col_idx == 1:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 2:
            c.font = Font(name='Arial', size=9, bold=True)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col_idx == 3:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col_idx == 4:
            c.font = Font(name='Arial', size=9, color='1F497D')
            c.fill = PatternFill('solid', fgColor='EBF3FB')
        elif col_idx == 5:
            c.font = Font(name='Arial', size=9, bold=True, color='1F7A1F')
            c.fill = PatternFill('solid', fgColor='C6EFCE')
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 6:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
        else:
            c.font = Font(name='Arial', size=8, color='595959')
            c.fill = PatternFill('solid', fgColor=bg)
    ws1.row_dimensions[row_idx].height = 50

ws1.freeze_panes = 'A3'

# =============== Sheet 2: 子维度 -> 爱好 ===============
ws2 = wb.create_sheet('子维度提升爱好映射')
ws2.sheet_view.showGridLines = False

ws2['A1'] = '27个子维度 - 能够提升该维度的兴趣爱好推荐'
ws2['A1'].font = Font(name='Arial', bold=True, size=13)
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

h2 = ['子维度', '大类', '维度说明', '能提升该维度的爱好（核心）', '典型提升方式']
for col, text in enumerate(h2, 1):
    c = ws2.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 32
ws2.column_dimensions['D'].width = 36
ws2.column_dimensions['E'].width = 40

for row_idx, dim in enumerate(dims_full, 3):
    cat = cat_map.get(dim, '其他')
    bg = cat_colors.get(cat, 'FFFFFF')
    hobby_list = dim_hobbies.get(dim, [])
    dim_hobbies_list = ['、'.join(dim_hobbies[dim])] if dim_hobbies[dim] else ['暂无数据']

    c = ws2.cell(row=row_idx, column=1, value=dim)
    c.font = Font(name='Arial', size=9, bold=True)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = border

    c = ws2.cell(row=row_idx, column=2, value=cat)
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='top')
    c.border = border

    c = ws2.cell(row=row_idx, column=3, value=dim_defs.get(dim, ''))
    c.font = Font(name='Arial', size=8)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    c = ws2.cell(row=row_idx, column=4, value='、'.join(hobby_list) if hobby_list else '暂无数据')
    c.font = Font(name='Arial', size=9, color='1F497D')
    c.fill = PatternFill('solid', fgColor='EBF3FB')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    # Get growth method from first hobby that boosts this dim
    methods = []
    for h in hobbies:
        if dim in h['core_dims']:
            methods.append('【' + h['name'] + '】' + h['growth'][:35])
            if len(methods) >= 3:
                break
    c = ws2.cell(row=row_idx, column=5, value='\n'.join(methods) if methods else '暂无数据')
    c.font = Font(name='Arial', size=8)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    ws2.row_dimensions[row_idx].height = 70

ws2.freeze_panes = 'A3'

# =============== Sheet 3: 矩阵 ===============
ws3 = wb.create_sheet('提升关系矩阵')
ws3.sheet_view.showGridLines = False

ws3['A1'] = '42个兴趣爱好 x 27个子维度 - 提升关系矩阵'
ws3['A1'].font = Font(name='Arial', bold=True, size=12)
ws3.merge_cells('A1:AJ1')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 12
for i in range(3, 30):
    ws3.column_dimensions[get_column_letter(i)].width = 8

for col, text in enumerate(['兴趣爱好', '大类'], 1):
    c = ws3.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

for col, dim in enumerate(dims_full, 3):
    sname = short_map.get(dim, dim)
    c = ws3.cell(row=2, column=col, value=sname)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws3.row_dimensions[2].height = 48

for row_idx, h in enumerate(hobbies, 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'

    c = ws3.cell(row=row_idx, column=1, value=h['name'])
    c.font = Font(name='Arial', size=8, bold=True)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = border

    c = ws3.cell(row=row_idx, column=2, value=h['cat'])
    c.font = Font(name='Arial', size=8)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

    for col_idx, dim in enumerate(dims_full, 3):
        if dim in h['core_dims']:
            c = ws3.cell(row=row_idx, column=col_idx, value='★')
            c.font = Font(name='Arial', size=9, color='1F7A1F', bold=True)
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        else:
            c = ws3.cell(row=row_idx, column=col_idx, value='')
            c.fill = PatternFill('solid', fgColor='F5F5F5')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws3.row_dimensions[row_idx].height = 22

# Legend
lr = len(hobbies) + 5
ws3.cell(row=lr, column=1, value='图例').font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=lr+1, column=1, value='★').font = Font(name='Arial', size=12, color='1F7A1F', bold=True)
ws3.cell(row=lr+1, column=1).fill = PatternFill('solid', fgColor='C6EFCE')
ws3.cell(row=lr+1, column=1).alignment = Alignment(horizontal='center')
ws3.cell(row=lr+1, column=2, value='核心提升：该兴趣爱好可显著提升此维度').font = Font(name='Arial', size=9)
ws3.merge_cells(f'B{lr+1}:H{lr+1}')
ws3.cell(row=lr+2, column=2, value='空白：该兴趣爱好与该维度关联较弱，不作为主要提升路径').font = Font(name='Arial', size=9)

ws3.cell(row=lr+4, column=1, value='大类颜色').font = Font(name='Arial', bold=True, size=10)
for i, (cn, cl) in enumerate(cat_colors.items()):
    r = lr + 5 + i
    ws3.cell(row=r, column=1, value=cn).font = Font(name='Arial', size=9)
    ws3.cell(row=r, column=1).fill = PatternFill('solid', fgColor=cl)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal='center')

ws3.freeze_panes = 'C3'

# =============== Sheet 4: 子维度定义 ===============
ws4 = wb.create_sheet('子维度定义与爱好')
ws4.sheet_view.showGridLines = False
ws4['A1'] = '27个子维度定义与可提升的爱好'
ws4['A1'].font = Font(name='Arial', bold=True, size=13)
ws4.merge_cells('A1:D1')
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 28

for col, text in enumerate(['子维度', '大类', '维度定义', '能提升该维度的爱好'], 1):
    c = ws4.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 36

for row_idx, dim in enumerate(dims_full, 3):
    cat = cat_map.get(dim, '其他')
    bg = cat_colors.get(cat, 'FFFFFF')

    c = ws4.cell(row=row_idx, column=1, value=dim)
    c.font = Font(name='Arial', size=9, bold=True)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = border

    c = ws4.cell(row=row_idx, column=2, value=cat)
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='top')
    c.border = border

    c = ws4.cell(row=row_idx, column=3, value=dim_defs.get(dim, ''))
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    hobby_list = dim_hobbies.get(dim, [])
    c = ws4.cell(row=row_idx, column=4, value='、'.join(hobby_list) if hobby_list else '暂无数据')
    c.font = Font(name='Arial', size=9, color='1F497D')
    c.fill = PatternFill('solid', fgColor='EBF3FB')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border

    ws4.row_dimensions[row_idx].height = 55

ws4.freeze_panes = 'A3'

output_path = r'C:\Users\25890\Desktop\兴趣爱好\兴趣爱好提升维度.xlsx'
wb.save(output_path)
print('Done! Saved to:', output_path)
print('Hobbies:', len(hobbies))
for dim in dims_full:
    count = len(dim_hobbies.get(dim, []))
    if count > 0:
        print(f'  {dim}: {count} hobbies -> {dim_hobbies[dim]}')
