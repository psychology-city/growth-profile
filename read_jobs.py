# -*- coding: utf-8 -*-
import json, sys
from openpyxl import load_workbook

try:
    wb = load_workbook(r'C:\Users\25890\Desktop\工作需求表_个人成长名片.xlsx')
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print(f"Loaded {len(rows)} rows from work Excel")
    print("Headers:", rows[1] if len(rows) > 1 else "No headers")
    if len(rows) > 2:
        print("Sample row 3:", rows[2])
        print("Sample row 4:", rows[3])
    sys.exit(0)
except Exception as e:
    print(f"Error: {e}")
    sys.exit(1)
