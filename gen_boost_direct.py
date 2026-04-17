# -*- coding: utf-8 -*-
import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read detail data
with open(r'C:\Users\25890\.qclaw\workspace\growth-profile\detail_data.json', 'r', encoding='utf-8') as f:
    detail_raw = json.load(f)

wb = Workbook()
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Sub-dimension definitions
dim_defs = {
    '注意力': '指长时间保持专注、不被干扰分散的能力，是深度工作的核心前提',
    '思维与洞察力': '透过现象看本质、快速抓住核心规律的能力，是问题分析和决策判断的基础',
    '好奇心与学习力': '主动探索未知、持续学习的内驱力，是知识积累和能力迭代的核心燃料',
    '信息分析与战略思维': '对复杂信息进行结构化分析、形成全局判断的能力，是战略决策的认知基础',
    '元认知能力': '对自己思维过程的监控和调节能力，是自我迭代和高效学习的元能力',
    '知识体系与跨界迁移': '构建系统知识框架、将一个领域的知识迁移到其他领域的能力',
    '开放判断力': '接纳不同观点、容纳矛盾信息、不被偏见局限的能力，是多元思维的基础',
    '创造力与创新': '产生新颖且有价值想法的能力，是产品创新和内容创作的核心驱动',
    '实践落地与反馈迭代': '将想法转化为行动、基于反馈快速优化的能力，是从"知道"到"做到"的关键',
    '共情内核与包容力': '感知和理解他人情绪与立场的能力，是人际沟通和客户服务的情感基础',
    '关系构建与维护': '主动拓展和维护有价值关系的能力，是资源整合和协作网络的核心',
    '协作沟通与社交智慧': '在多人协作中有效表达，协调分歧、达成共识的能力',
    '赋能型领导力': '通过支持他人成长来实现团队目标的能力，是从个人贡献者到管理者的关键跃迁',
    '内在信念与品格': '在压力和诱惑下坚守核心价值观的能力，是长期信任和人品的基石',
    '目标推进与执行力': '将目标转化为持续行动、克服阻力拿到结果的能力',
    '自我管控与风险平衡': '控制冲动，管理情绪、平衡风险的能力，是成熟决策的行为保障',
    '决策引领与变革': '在不确定性中做出果断决策并推动落地的能力，是管理者的核心职责',
    '价值交付型服务能力': '以客户需求为中心、交付超出预期的价值的意识和行为模式',
    '前瞻引领型领导力': '为团队指明方向、凝聚人心、引领变革的领导特质',
    '落地统筹型执行能力': '将模糊目标拆解为可执行动作，协调多方资源、确保落地的综合能力',
    '魅力调节因子': '通过真诚、幽默、感恩等社交魅力放大其他能力落地效果的能力',
    '自我超越意愿': '不满足于现状、主动突破舒适区、持续追求成长的内驱力',
    '人生意义建构': '赋予工作和生活超越功利目标的意义感，是持久动力的深层来源',
    '利他与公共贡献': '愿意为他人、行业、社会创造价值的意识，是领导力和影响力的社会基础',
    '超越性信念': '相信宏大规律和长期价值、愿意承受短期不确定性的信念感',
    '欣赏美与卓越': '感知美、追求卓越、欣赏高品质事物的意识，是审美和创意的基础',
    '超越性感恩与宽恕': '对已拥有的一切心怀感恩、对过去释怀放下的心态，是情绪稳定的深层来源',
}

# Sub-dimension short names
dim_short = {
    '注意力': '注意力', '思维与洞察力': '思维洞察', '好奇心与学习力': '好奇心',
    '信息分析与战略思维': '信息分析', '元认知能力': '元认知',
    '知识体系与跨界迁移': '知识迁移', '开放判断力': '开放判断',
    '创造力与创新': '创造力', '实践落地与反馈迭代': '实践落地',
    '共情内核与包容力': '共情内核', '关系构建与维护': '关系构建',
    '协作沟通与社交智慧': '协作沟通', '赋能型领导力': '赋能领导',
    '内在信念与品格': '内在信念', '目标推进与执行力': '目标推进',
    '自我管控与风险平衡': '自我管控', '决策引领与变革': '决策引领',
    '价值交付型服务能力': '价值交付', '前瞻引领型领导力': '前瞻领导',
    '落地统筹型执行能力': '落地统筹', '魅力调节因子': '魅力调节',
    '自我超越意愿': '自我超越', '人生意义建构': '意义建构',
    '利他与公共贡献': '利他贡献', '超越性信念': '超越信念',
    '欣赏美与卓越': '欣赏美', '超越性感恩与宽恕': '超越感恩',
}

# Sub-dimension categories
dim_cat = {
    '注意力': '认知', '思维与洞察力': '认知', '好奇心与学习力': '认知',
    '信息分析与战略思维': '认知', '元认知能力': '认知',
    '知识体系与跨界迁移': '认知', '开放判断力': '认知',
    '创造力与创新': '认知', '实践落地与反馈迭代': '认知',
    '共情内核与包容力': '人际', '关系构建与维护': '人际',
    '协作沟通与社交智慧': '人际', '赋能型领导力': '人际',
    '内在信念与品格': '意志', '目标推进与执行力': '意志',
    '自我管控与风险平衡': '意志', '决策引领与变革': '意志',
    '价值交付型服务能力': '领导', '前瞻引领型领导力': '领导',
    '落地统筹型执行能力': '领导', '魅力调节因子': '领导',
    '自我超越意愿': '超越', '人生意义建构': '超越',
    '利他与公共贡献': '超越', '超越性信念': '超越',
    '欣赏美与卓越': '超越', '超越性感恩与宽恕': '超越',
}

cat_colors = {
    '认知': 'DAE8FC', '人际': 'D5E8D4',
    '意志': 'FFF2CC', '领导': 'F8CECC', '超越': 'E1D5E7',
}

dim_order = [
    '注意力', '思维与洞察力', '好奇心与学习力', '信息分析与战略思维',
    '元认知能力', '知识体系与跨界迁移', '开放判断力', '创造力与创新',
    '实践落地与反馈迭代', '共情内核与包容力', '关系构建与维护',
    '协作沟通与社交智慧', '赋能型领导力', '内在信念与品格',
    '目标推进与执行力', '自我管控与风险平衡', '决策引领与变革',
    '价值交付型服务能力', '前瞻引领型领导力', '落地统筹型执行能力',
    '魅力调节因子', '自我超越意愿', '人生意义建构',
    '利他与公共贡献', '超越性信念', '欣赏美与卓越', '超越性感恩与宽恕',
]

# Hobby boost data - (hobby, cat, boost_dims, level, method)
hobby_boost = [
    ('跑步', '运动健身', ['注意力', '内在信念与品格', '意志力', '自我管控与风险平衡', '超越性感恩与宽恕'], '核心提升', '长跑需要持续专注克服疲劳；跑完后成就感培养感恩心态；长期坚持培养内在信念'),
    ('健身', '运动健身', ['目标推进与执行力', '意志力', '赋能型领导力', '内在信念与品格', '自我超越意愿'], '核心提升', '系统训练需要设定目标并持续推进；重量突破培养抗压；带别人一起训练培养赋能领导力'),
    ('游泳', '运动健身', ['注意力', '自我管控与风险平衡', '内在信念与品格', '意志力'], '核心提升', '水中运动需要极强的注意力控制；长距离游泳培养意志力和内在信念'),
    ('球类运动', '运动健身', ['协作沟通与社交智慧', '共情内核与包容力', '目标推进与执行力', '魅力调节因子'], '核心提升', '团队运动需要实时感知队友意图（共情）、高效沟通协作；比赛中的魅力展现培养领袖气质'),
    ('骑行', '运动健身', ['目标推进与执行力', '内在信念与品格', '自我超越意愿', '意志力'], '核心提升', '长途骑行目标明确；面对坡道和恶劣天气培养抗压；突破个人极限培养自我超越'),
    ('格斗训练', '运动健身', ['内在信念与品格', '意志力', '目标推进与执行力', '魅力调节因子'], '核心提升', '对抗训练在高度压力下保持冷静；实战中培养内在信念；擂台表现培养舞台魅力'),
    ('瑜伽冥想', '运动健身', ['注意力', '元认知能力', '自我管控与风险平衡', '超越性信念'], '核心提升', '冥想直接训练注意力肌肉；正念练习大幅提升元认知；体式练习培养内在信念'),
    ('极限运动', '运动健身', ['超越性信念', '自我超越意愿', '意志力', '超越性感恩与宽恕'], '核心提升', '极限挑战需要超越对恐惧的信念；高海拔深海等极端环境培养超越感恩'),
    ('乐器演奏', '音乐艺术', ['注意力', '目标推进与执行力', '意志力', '自我超越意愿'], '核心提升', '乐器练习是注意力的高强度训练；考级目标推进意志力；技巧突破培养自我超越'),
    ('唱歌声乐', '音乐艺术', ['共情内核与包容力', '魅力调节因子', '协作沟通与社交智慧', '开放判断力'], '核心提升', '歌曲演唱需要共情歌词情感；舞台表演培养魅力；合唱需要协调沟通'),
    ('绘画艺术', '音乐艺术', ['创造力与创新', '欣赏美与卓越', '注意力', '好奇心与学习力', '实践落地与反馈迭代'], '核心提升', '创作过程直接培养创造力；观察和记录培养审美；反复修改培养实践落地能力'),
    ('书法', '音乐艺术', ['注意力', '内在信念与品格', '超越性信念', '目标推进与执行力'], '核心提升', '书法需要极度专注；长期修炼培养内在信念；草书等高阶需要超越性信念'),
    ('摄影', '音乐艺术', ['欣赏美与卓越', '开放判断力', '好奇心与学习力', '创造力与创新'], '核心提升', '拍摄需要发现美的眼睛（审美）；不同题材拓展认知边界（开放判断）；创意表达培养创造力'),
    ('舞蹈', '音乐艺术', ['共情内核与包容力', '欣赏美与卓越', '创造力与创新', '协作沟通与社交智慧', '魅力调节因子'], '核心提升', '舞蹈表达需要共情音乐情感；编舞培养创造力；团队舞蹈培养协作和魅力'),
    ('手工艺', '音乐艺术', ['实践落地与反馈迭代', '欣赏美与卓越', '好奇心与学习力', '创造力与创新'], '核心提升', '动手创作直接训练实践落地；制作过程培养审美和好奇心；原创设计培养创造力'),
    ('阅读', '知识学习', ['元认知能力', '好奇心与学习力', '内在信念与品格', '注意力', '自我超越意愿'], '核心提升', '深度阅读训练元认知；不同书籍拓展认知边界；长期阅读培养内在信念和自我超越'),
    ('写作创作', '知识学习', ['创造力与创新', '好奇心与学习力', '元认知能力', '自我超越意愿', '人生意义建构'], '核心提升', '写作是思想的重组和创造；持续输出倒逼输入培养好奇心；长篇写作培养意义建构'),
    ('语言学习', '知识学习', ['好奇心与学习力', '开放判断力', '元认知能力', '自我超越意愿', '目标推进与执行力'], '核心提升', '学习新语言需要突破舒适区（自我超越）；理解不同文化培养开放判断力；语法学习训练元认知'),
    ('口播演讲', '知识学习', ['开放判断力', '共情内核与包容力', '魅力调节因子', '协作沟通与社交智慧'], '核心提升', '需要站在听众角度思考（共情）；表达观点培养开放判断力；舞台表现培养魅力'),
    ('电子游戏', '游戏娱乐', ['信息分析与战略思维', '注意力', '创造力与创新', '好奇心与学习力', '目标推进与执行力'], '核心提升', '策略游戏训练战略思维；竞技游戏高强度训练注意力；沙盒游戏培养创造力和好奇心'),
    ('桌游剧本杀', '游戏娱乐', ['协作沟通与社交智慧', '开放判断力', '共情内核与包容力', '魅力调节因子'], '核心提升', '多人博弈需要协作沟通和开放判断；剧本杀需要共情角色；主持控场培养魅力'),
    ('棋类', '游戏娱乐', ['注意力', '信息分析与战略思维', '元认知能力', '自我管控与风险平衡', '目标推进与执行力'], '核心提升', '对弈需要高度专注；复盘训练元认知；长局训练自我管控和目标推进'),
    ('旅行出游', '户外探索', ['开放判断力', '好奇心与学习力', '共情内核与包容力', '超越性信念', '人生意义建构'], '核心提升', '接触不同文化培养开放判断力；探索未知培养好奇心；见世界培养超越信念和人生意义'),
    ('徒步露营', '户外探索', ['意志力', '内在信念与品格', '意志力', '超越性信念', '协作沟通与社交智慧'], '核心提升', '长距离徒步培养意志力和抗压；野外露营培养超越信念；团队出行培养协作沟通'),
    ('观鸟自然观察', '户外探索', ['注意力', '好奇心与学习力', '开放判断力', '欣赏美与卓越'], '核心提升', '观鸟需要长时间安静专注；识别物种培养开放判断力；发现美培养审美感知'),
    ('钓鱼', '户外探索', ['注意力', '自我管控与风险平衡', '超越性感恩与宽恕', '内在信念与品格'], '核心提升', '长时间守候培养注意力和自我管控；接受空军培养超越感恩；磨练心性培养内在信念'),
    ('烹饪烘焙', '生活美食', ['实践落地与反馈迭代', '欣赏美与卓越', '创造力与创新', '好奇心与学习力'], '核心提升', '做菜需要动手实践落地；摆盘培养审美；创新菜谱培养创造力和好奇心'),
    ('品酒茶道', '生活美食', ['欣赏美与卓越', '超越性信念', '共情内核与包容力', '内在信念与品格'], '核心提升', '品鉴需要精细的审美感知；茶道培养超越信念；品鉴交流培养共情和内在信念'),
    ('园艺绿植', '生活美食', ['超越性感恩与宽恕', '注意力', '内在信念与品格', '自我超越意愿'], '核心提升', '植物成长需要耐心等待（超越感恩）；养护过程培养注意力和内在信念；收获培养自我超越'),
    ('收藏', '生活美食', ['欣赏美与卓越', '好奇心与学习力', '目标推进与执行力', '开放判断力'], '核心提升', '系统收藏需要审美眼光；研究价值培养好奇心；制定收藏目标培养目标推进'),
    ('编程开发', '科技数码', ['注意力', '信息分析与战略思维', '目标推进与执行力', '实践落地与反馈迭代', '元认知能力'], '核心提升', '编程是注意力的高强度训练；逻辑推理是核心能力；项目开发培养目标推进和实践落地'),
    ('AI工具使用', '科技数码', ['好奇心与学习力', '开放判断力', '实践落地与反馈迭代', '欣赏美与卓越', '创造力与创新'], '核心提升', 'AI快速迭代需要持续好奇心；尝试不同工具培养开放判断力；AI创作培养创造力和审美'),
    ('硬件极客DIY', '科技数码', ['实践落地与反馈迭代', '好奇心与学习力', '目标推进与执行力', '注意力'], '核心提升', '动手拆装直接训练实践落地；研究原理培养好奇心；项目目标培养注意力'),
    ('无人机航拍', '科技数码', ['注意力', '目标推进与执行力', '欣赏美与卓越', '意志力'], '核心提升', '飞行操控需要高度专注；拍摄创作培养审美感知；应对突发状况培养抗压能力'),
    ('志愿服务', '社交公益', ['共情内核与包容力', '利他与公共贡献', '协作沟通与社交智慧', '超越性感恩与宽恕'], '核心提升', '志愿服务直接培养共情和利他意愿；组织活动培养协作沟通；帮助他人培养超越感恩'),
    ('社群运营', '社交公益', ['协作沟通与社交智慧', '赋能型领导力', '魅力调节因子', '共情内核与包容力'], '核心提升', '社群运营需要多方协调培养协作沟通；赋能成员培养领导力；真诚互动培养魅力'),
    ('辩论演讲', '社交公益', ['开放判断力', '协作沟通与社交智慧', '魅力调节因子', '共情内核与包容力'], '核心提升', '辩论需要从对方角度思考培养共情和开放判断力；舞台表现培养魅力；说服他人培养协作沟通'),
    ('心理疗愈', '社交公益', ['共情内核与包容力', '元认知能力', '超越性感恩与宽恕', '人生意义建构'], '核心提升', '自我探索大幅提升元认知；学会感恩培养超越感恩；理解人性培养共情和人生意义建构'),
    ('金融投资', '知识学习', ['信息分析与战略思维', '意志力', '自我管控与风险平衡', '超越性信念'], '核心提升', '投资分析训练信息分析能力；市场波动培养抗压；长期价值投资培养超越信念'),
    ('历史哲学', '知识学习', ['好奇心与学习力', '元认知能力', '内在信念与品格', '超越性信念', '开放判断力'], '核心提升', '历史拓展认知边界培养好奇心；哲学思考训练元认知；思想史培养超越信念和内在信念'),
    ('科学实验', '知识学习', ['好奇心与学习力', '信息分析与战略思维', '实践落地与反馈迭代', '注意力', '元认知能力'], '核心提升', '实验探究直接培养好奇心和逻辑推理；动手操作训练实践落地；观察记录培养注意力和元认知'),
    ('考证学习', '知识学习', ['目标推进与执行力', '意志力', '自我管控与风险平衡', '意志力', '内在信念与品格'], '核心提升', '备考需要设定目标并持续推进；考前压力培养抗压；坚持学习培养意志力和自我管控'),
]

print(f'Hobby boost entries: {len(hobby_boost)}')

# ============ Sheet 1: Summary Table ============
ws1 = wb.active
ws1.title = '爱好提升维度总表'
ws1.sheet_view.showGridLines = False

ws1['A1'] = '兴趣爱好能够提升的子维度总览'
ws1['A1'].font = Font(name='Arial', bold=True, size=13)
ws1.merge_cells('A1:G1')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 28

headers1 = ['序号', '兴趣爱好', '大类', '能提升的子维度', '提升等级', '提升方式说明']
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

col_widths1 = [6, 16, 12, 40, 10, 46]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for row_idx, (hobby, cat, dims, level, method) in enumerate(hobby_boost, 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'

    for col, val in [(1, row_idx-2), (2, hobby), (3, cat), (4, '、'.join(dims)), (5, level), (6, method)]:
        c = ws1.cell(row=row_idx, column=col, value=val)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if col == 1:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col == 2:
            c.font = Font(name='Arial', size=9, bold=True)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col == 3:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
        elif col == 4:
            c.font = Font(name='Arial', size=9, color='1F497D')
            c.fill = PatternFill('solid', fgColor='EBF3FB')
        elif col == 5:
            c.font = Font(name='Arial', size=9, bold=True)
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        else:
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
        if col == 1:
            c.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[row_idx].height = 42

ws1.freeze_panes = 'A3'

# ============ Sheet 2: Dimension -> Hobby Mapping ============
ws2 = wb.create_sheet('子维度提升爱好映射')
ws2.sheet_view.showGridLines = False

ws2['A1'] = '27个子维度 - 能够提升该维度的兴趣爱好推荐'
ws2['A1'].font = Font(name='Arial', bold=True, size=13)
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

# Build mapping
dim_to_hobbies = {dim: [] for dim in dim_order}
for hobby, cat, dims, level, method in hobby_boost:
    for dim in dims:
        dim_to_hobbies[dim].append(hobby)

ws2.cell(row=2, column=1, value='子维度').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws2.cell(row=2, column=1).fill = PatternFill('solid', fgColor='2F5496')
ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=2, column=1).border = border

ws2.cell(row=2, column=2, value='大类').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws2.cell(row=2, column=2).fill = PatternFill('solid', fgColor='2F5496')
ws2.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=2, column=2).border = border

ws2.cell(row=2, column=3, value='维度说明').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws2.cell(row=2, column=3).fill = PatternFill('solid', fgColor='2F5496')
ws2.cell(row=2, column=3).alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=2, column=3).border = border

ws2.cell(row=2, column=4, value='可提升该维度的兴趣爱好').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws2.cell(row=2, column=4).fill = PatternFill('solid', fgColor='2F5496')
ws2.cell(row=2, column=4).alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=2, column=4).border = border

ws2.cell(row=2, column=5, value='推荐理由').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws2.cell(row=2, column=5).fill = PatternFill('solid', fgColor='2F5496')
ws2.cell(row=2, column=5).alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=2, column=5).border = border

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 36
ws2.column_dimensions['E'].width = 40

for row_idx, dim in enumerate(dim_order, 3):
    cat = dim_cat[dim]
    bg = cat_colors.get(cat, 'FFFFFF')
    hobbies_list = dim_to_hobbies[dim]

    for col, val in [(1, dim), (2, cat), (3, dim_defs.get(dim, '')), (4, '、'.join(hobbies_list) if hobbies_list else '暂无数据')]:
        c = ws2.cell(row=row_idx, column=col, value=val)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.fill = PatternFill('solid', fgColor=bg)
        if col == 1:
            c.font = Font(name='Arial', size=9, bold=True)
        elif col == 4:
            c.font = Font(name='Arial', size=9, color='1F497D')
        else:
            c.font = Font(name='Arial', size=8)

    # Get method for this dim from first hobby that boosts it
    methods = []
    for hobby, cat2, dims2, level, method in hobby_boost:
        if dim in dims2:
            methods.append(method[:40])
            if len(methods) >= 3:
                break
    c = ws2.cell(row=row_idx, column=5, value='\n'.join(methods))
    c.font = Font(name='Arial', size=8)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = border
    c.fill = PatternFill('solid', fgColor='FFFFFF')

    ws2.row_dimensions[row_idx].height = 60

ws2.freeze_panes = 'A3'

# ============ Sheet 3: Matrix ============
ws3 = wb.create_sheet('提升关系矩阵')
ws3.sheet_view.showGridLines = False

ws3['A1'] = '42个兴趣爱好 x 27个子维度 - 提升关系矩阵'
ws3['A1'].font = Font(name='Arial', bold=True, size=12)
ws3.merge_cells('A1:AJ1')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 12
for i in range(3, 30):
    ws3.column_dimensions[get_column_letter(i)].width = 8

for col, text in enumerate(['兴趣爱好', '大类'], 1):
    c = ws3.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

for col, dim in enumerate(dim_order, 3):
    short = dim_short[dim]
    c = ws3.cell(row=2, column=col, value=short)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws3.row_dimensions[2].height = 45

hobby_names = [h[0] for h in hobby_boost]
hobby_cats = [h[1] for h in hobby_boost]
hobby_dims_list = [h[2] for h in hobby_boost]

for row_idx, (h_name, h_cat, h_dims) in enumerate(zip(hobby_names, hobby_cats, hobby_dims_list), 3):
    bg = row_idx % 2 == 0 and 'F2F7FF' or 'FFFFFF'

    c = ws3.cell(row=row_idx, column=1, value=h_name)
    c.font = Font(name='Arial', size=8, bold=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = border
    c.fill = PatternFill('solid', fgColor=bg)

    c = ws3.cell(row=row_idx, column=2, value=h_cat)
    c.font = Font(name='Arial', size=8)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    c.fill = PatternFill('solid', fgColor=bg)

    for col_idx, dim in enumerate(dim_order, 3):
        if dim in h_dims:
            c = ws3.cell(row=row_idx, column=col_idx, value='★')
            c.font = Font(name='Arial', size=9, color='1F7A1F', bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        else:
            c = ws3.cell(row=row_idx, column=col_idx, value='')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
            c.fill = PatternFill('solid', fgColor='F5F5F5')
    ws3.row_dimensions[row_idx].height = 22

# Legend
lr = len(hobby_names) + 5
ws3.cell(row=lr, column=1, value='图例').font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=lr+1, column=1, value='★').font = Font(name='Arial', size=12, color='1F7A1F', bold=True)
ws3.cell(row=lr+1, column=1).fill = PatternFill('solid', fgColor='C6EFCE')
ws3.cell(row=lr+1, column=1).alignment = Alignment(horizontal='center')
ws3.cell(row=lr+1, column=2, value='该兴趣爱好可显著提升此维度（成长型高分时更适配该爱好；等级型高分者更需通过该爱好补足）').font = Font(name='Arial', size=9)
ws3.merge_cells(f'B{lr+1}:H{lr+1}')
ws3.cell(row=lr+2, column=2, value='空白=该兴趣爱好与该维度关联较弱').font = Font(name='Arial', size=9)

ws3.cell(row=lr+4, column=1, value='大类颜色').font = Font(name='Arial', bold=True, size=10)
for i, (cat_name, color) in enumerate(cat_colors.items()):
    r = lr + 5 + i
    ws3.cell(row=r, column=1, value=cat_name).font = Font(name='Arial', size=9)
    ws3.cell(row=r, column=1).fill = PatternFill('solid', fgColor=color)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal='center')

ws3.freeze_panes = 'C3'

# ============ Sheet 4: Dimension descriptions ============
ws4 = wb.create_sheet('子维度定义')
ws4.sheet_view.showGridLines = False
ws4['A1'] = '27个子维度定义与说明'
ws4['A1'].font = Font(name='Arial', bold=True, size=13)
ws4.merge_cells('A1:D1')
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 28

for col, h in enumerate(['子维度', '所属大类', '维度定义', '能提升该维度的爱好'], 1):
    c = ws4.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 42
ws4.column_dimensions['D'].width = 36

for row_idx, dim in enumerate(dim_order, 3):
    cat = dim_cat[dim]
    bg = cat_colors.get(cat, 'FFFFFF')
    hobbies_list = dim_to_hobbies.get(dim, [])

    for col, val in [(1, dim), (2, cat), (3, dim_defs.get(dim, '')), (4, '、'.join(hobbies_list) if hobbies_list else '暂无数据')]:
        c = ws4.cell(row=row_idx, column=col, value=val)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.fill = PatternFill('solid', fgColor=bg)
        if col == 1:
            c.font = Font(name='Arial', size=9, bold=True)
        elif col == 4:
            c.font = Font(name='Arial', size=9, color='1F497D')
        else:
            c.font = Font(name='Arial', size=9)
    ws4.row_dimensions[row_idx].height = 50

ws4.freeze_panes = 'A3'

output_path = r'C:\Users\25890\Desktop\兴趣爱好\兴趣爱好提升维度.xlsx'
wb.save(output_path)
print(f'Successfully saved: {output_path}')
