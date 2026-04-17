# 题目影响分析表生成脚本
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============ Sheet 1: 说明 ============
ws_info = wb.active
ws_info.title = "说明与评分标准"

info_font = Font(name='Arial', size=10)
title_font = Font(name='Arial', bold=True, size=13)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', fgColor='4F81BD')
body_font = Font(name='Arial', size=9)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
influence_fill_yes = PatternFill('solid', fgColor='C6EFCE')   # 浅绿 = 有影响
influence_fill_no  = PatternFill('solid', fgColor='FFCCCC')   # 浅红 = 无影响
influence_fill_mid = PatternFill('solid', fgColor='FFEB9C')   # 浅黄 = 中等影响

# 说明内容
ws_info['A1'] = '题目与工作匹配度影响分析表 — 说明'
ws_info['A1'].font = title_font
ws_info.merge_cells('A1:D1')

ws_info['A3'] = '■ 评分说明'
ws_info['A3'].font = Font(name='Arial', bold=True, size=11)
ws_info['A4'] = '每道题目（1-5分）都从两个维度测评：'
ws_info['A5'] = '  · 等级模式题：高分（4-5分）代表该能力维度较弱/有卡点'
ws_info['A6'] = '  · 成长模式题：高分（4-5分）代表该能力维度较强/有优势'
ws_info['A7'] = '  · 最终得分 = 成长模式总分 - 等级模式总分（差值越大，成长型特征越明显）'

ws_info['A9'] = '■ 影响程度定义'
items = [
    ['影响程度', '含义', '填充色'],
    ['★★★★★ 强影响', '选择该选项的得分，直接决定是否能胜任该工作的核心要求', '浅绿'],
    ['★★★☆☆ 中影响', '该维度对工作有正向帮助，得高分有助于提升工作表现', '浅黄'],
    ['☆☆☆☆☆ 无影响', '该维度在该工作中不是关键能力，高低分均不影响胜任力', '浅红'],
]
ws_info['A10'] = items[0][0]; ws_info['A10'].font = Font(bold=True)
ws_info['B10'] = items[0][1]; ws_info['B10'].font = Font(bold=True)
ws_info['C10'] = items[0][2]; ws_info['C10'].font = Font(bold=True)
for item in items[1:]:
    ws_info.append(item)

ws_info['A15'] = '■ 典型Likert量表选项参考'
ws_info['A15'].font = Font(name='Arial', bold=True, size=11)
opt_headers = ['分值', '等级模式典型描述（高分=能力弱）', '成长模式典型描述（高分=能力强）']
for col, h in enumerate(opt_headers, 1):
    c = ws_info.cell(row=16, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.border = border
opts = [
    ['1分', '完全不符合该描述（说明能力正常）', '完全符合该描述（说明能力强）'],
    ['2分', '比较不符合（轻微卡点）', '比较符合（较强能力）'],
    ['3分', '一般/不确定', '一般/不确定'],
    ['4分', '比较符合（明显卡点）', '比较符合（很强能力）'],
    ['5分', '完全符合（严重卡点）', '完全符合（极强能力）'],
]
for row_data in opts:
    row_num = ws_info.max_row + 1
    for col, val in enumerate(row_data, 1):
        c = ws_info.cell(row=row_num, column=col, value=val)
        c.font = body_font; c.border = border

ws_info.column_dimensions['A'].width = 25
ws_info.column_dimensions['B'].width = 40
ws_info.column_dimensions['C'].width = 40

# ============ Sheet 2+: 各子维度分析 ============
# 定义27个子维度的分析数据
dimensions = [
    {
        'id': 'cognitive_01', 'name': '注意力与专注力', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['后端开发工程师', '算法工程师', '硬件工程师', '测试工程师', '数据分析师', '财务分析师', '审计专员']),
            ('★★★☆☆ 中影响', ['前端开发工程师', 'UI设计师', '产品经理', '电商运营', 'HRBP', '法务专员']),
            ('☆☆☆☆☆ 无影响', ['销售经理', '客户成功经理', 'BD商务拓展', '行政专员', '客服专员']),
        ]
    },
    {
        'id': 'cognitive_02', 'name': '思维与洞察力', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['产品经理', '创新研究员', '战略规划师', '管理咨询顾问', '商业分析师', '行业研究员']),
            ('★★★☆☆ 中影响', ['后端开发工程师', '算法工程师', '品牌设计师', 'HRBP', 'OD组织发展专员']),
            ('☆☆☆☆☆ 无影响', ['数据录入员', '流水线操作', '基础客服', '行政专员']),
        ]
    },
    {
        'id': 'cognitive_03', 'name': '好奇心与学习能力', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['算法工程师', '创新研究员', '内容创作者', '行业研究员', '投资分析师', '产品经理']),
            ('★★★☆☆ 中影响', ['后端开发工程师', '前端开发工程师', '数据分析师', '培训发展专员', 'HRBP']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服']),
        ]
    },
    {
        'id': 'cognitive_04', 'name': '信息分析与战略思维', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['战略规划师', '投资分析师', '行业研究员', '商业分析师', '数字营销专员', '财务分析师']),
            ('★★★☆☆ 中影响', ['产品经理', '品牌营销经理', '运营总监', 'HRBP', '管理咨询顾问']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '行政专员', '基础客服']),
        ]
    },
    {
        'id': 'cognitive_05', 'name': '元认知能力', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['产品经理', '管理咨询顾问', '项目经理', '运营总监', 'CEO/创业者']),
            ('★★★☆☆ 中影响', ['内容创作者', '算法工程师', '培训发展专员', 'HRBP']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '行政专员']),
        ]
    },
    {
        'id': 'cognitive_06', 'name': '知识体系与跨界迁移', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['创新研究员', '产品经理', '行业研究员', '培训发展专员', 'CEO/创业者']),
            ('★★★☆☆ 中影响', ['管理咨询顾问', '内容创作者', '品牌设计师', 'HRBP']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服']),
        ]
    },
    {
        'id': 'cognitive_07', 'name': '开放判断力', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['管理咨询顾问', 'HRBP', 'OD组织发展专员', '产品经理', 'HRD/CHO']),
            ('★★★☆☆ 中影响', ['战略规划师', '品牌营销经理', '培训发展专员', '投资分析师']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服']),
        ]
    },
    {
        'id': 'cognitive_08', 'name': '创造力与创新', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['产品经理', '创新研究员', 'UI设计师', '品牌设计师', '插画师', '内容创作者']),
            ('★★★☆☆ 中影响', ['前端开发工程师', '视频剪辑师', '品牌营销经理', '数字营销专员']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '审计专员']),
        ]
    },
    {
        'id': 'cognitive_09', 'name': '实践落地与反馈迭代', 'category': '认知能力',
        'jobs': [
            ('★★★★★ 强影响', ['项目经理', '运营总监', '电商运营', '销售经理', '产品经理']),
            ('★★★☆☆ 中影响', ['后端开发工程师', '前端开发工程师', '测试工程师', '品牌营销经理']),
            ('☆☆☆☆☆ 无影响', ['行政专员', '流水线操作']),
        ]
    },
    {
        'id': 'empathy_01', 'name': '共情内核与包容力', 'category': '人际共情',
        'jobs': [
            ('★★★★★ 强影响', ['HRBP', '客服专员', '客户成功经理', '培训发展专员', 'HRD/CHO', '产品经理']),
            ('★★★☆☆ 中影响', ['销售经理', '大客户经理', '项目经理', '运营总监', '行政专员']),
            ('☆☆☆☆☆ 无影响', ['算法工程师', '后端开发工程师', '硬件工程师', '数据录入员']),
        ]
    },
    {
        'id': 'empathy_02', 'name': '关系构建与维护', 'category': '人际共情',
        'jobs': [
            ('★★★★★ 强影响', ['销售经理', '大客户经理', 'BD商务拓展', '客户成功经理', 'HRBP']),
            ('★★★☆☆ 中影响', ['项目经理', '培训发展专员', '品牌营销经理', '内容运营']),
            ('☆☆☆☆☆ 无影响', ['算法工程师', '后端开发工程师', '硬件工程师', '数据录入员', '流水线操作']),
        ]
    },
    {
        'id': 'empathy_03', 'name': '协作沟通与社交智慧', 'category': '人际共情',
        'jobs': [
            ('★★★★★ 强影响', ['项目经理', '产品经理', 'HRBP', 'BD商务拓展', '管理咨询顾问', '运营总监']),
            ('★★★☆☆ 中影响', ['前端开发工程师', '市场销售类', '培训发展专员', '法务专员']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '硬件工程师']),
        ]
    },
    {
        'id': 'empathy_04', 'name': '赋能型领导力', 'category': '人际共情',
        'jobs': [
            ('★★★★★ 强影响', ['运营总监', '市场总监', '技术VP/CTO', 'HRD/CHO', 'CEO/创业者', '项目经理']),
            ('★★★☆☆ 中影响', ['HRBP', '产品经理', '销售经理', '培训发展专员']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '行政专员']),
        ]
    },
    {
        'id': 'will_01', 'name': '内在信念与品格内核', 'category': '意志行动',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '运营总监', '技术VP/CTO', '市场总监', '销售经理', '内容创作者']),
            ('★★★☆☆ 中影响', ['产品经理', '项目经理', '管理咨询顾问', '投资分析师', '算法工程师']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '行政专员']),
        ]
    },
    {
        'id': 'will_02', 'name': '目标推进与执行能力', 'category': '意志行动',
        'jobs': [
            ('★★★★★ 强影响', ['项目经理', '电商运营', '销售经理', '运营总监', '市场总监', '行政专员']),
            ('★★★☆☆ 中影响', ['产品经理', 'HRBP', '培训发展专员', '数据分析师', '前端开发工程师']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员']),
        ]
    },
    {
        'id': 'will_03', 'name': '自我管控与风险平衡', 'category': '意志行动',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '运营总监', '投资分析师', '财务分析师', '审计专员', '产品经理']),
            ('★★★☆☆ 中影响', ['项目经理', '销售经理', '管理咨询顾问', '技术VP/CTO', 'HRD/CHO']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '行政专员', '基础客服']),
        ]
    },
    {
        'id': 'will_04', 'name': '决策引领与变革能力', 'category': '意志行动',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '运营总监', '技术VP/CTO', '市场总监', 'HRD/CHO', '产品经理']),
            ('★★★☆☆ 中影响', ['项目经理', '战略规划师', '创新研究员', 'BD商务拓展', '销售经理']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '行政专员']),
        ]
    },
    {
        'id': 'leadership_01', 'name': '价值交付型服务能力', 'category': '领导职业',
        'jobs': [
            ('★★★★★ 强影响', ['客户成功经理', '客服专员', '大客户经理', '培训发展专员', '咨询顾问', '法务专员']),
            ('★★★☆☆ 中影响', ['产品经理', '销售经理', 'HRBP', '行政专员', '财务分析师']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '硬件工程师', '算法工程师', '数据录入员']),
        ]
    },
    {
        'id': 'leadership_02', 'name': '前瞻引领型领导力', 'category': '领导职业',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '运营总监', '市场总监', '技术VP/CTO', 'HRD/CHO', '战略规划师']),
            ('★★★☆☆ 中影响', ['产品经理', '创新研究员', '项目经理', '品牌营销经理']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '行政专员', '基础客服']),
        ]
    },
    {
        'id': 'leadership_03', 'name': '落地统筹型执行能力', 'category': '领导职业',
        'jobs': [
            ('★★★★★ 强影响', ['项目经理', '运营总监', '电商运营', '行政专员', 'HRBP', '市场总监']),
            ('★★★☆☆ 中影响', ['产品经理', '销售经理', '培训发展专员', '法务专员']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服']),
        ]
    },
    {
        'id': 'leadership_04', 'name': '魅力调节因子', 'category': '领导职业',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '运营总监', '市场总监', '销售经理', 'BD商务拓展', 'HRD/CHO']),
            ('★★★☆☆ 中影响', ['产品经理', '项目经理', 'HRBP', '培训发展专员', '内容创作者']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '硬件工程师', '算法工程师']),
        ]
    },
    {
        'id': 'transcend_01', 'name': '自我超越意愿', 'category': '超越动力',
        'jobs': [
            ('★★★★★ 强影响', ['算法工程师', '内容创作者', '创新研究员', 'CEO/创业者', '产品经理', '行业研究员']),
            ('★★★☆☆ 中影响', ['前端开发工程师', 'UI设计师', '战略规划师', '投资分析师', '运营总监']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '行政专员']),
        ]
    },
    {
        'id': 'transcend_02', 'name': '人生意义建构能力', 'category': '超越动力',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '内容创作者', '创新研究员', '产品经理', '行业研究员', '运营总监']),
            ('★★★☆☆ 中影响', ['产品经理', '培训发展专员', 'OD组织发展专员', '管理咨询顾问']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '行政专员']),
        ]
    },
    {
        'id': 'transcend_03', 'name': '利他与公共贡献意愿', 'category': '超越动力',
        'jobs': [
            ('★★★★★ 强影响', ['HRD/CHO', '培训发展专员', 'OD组织发展专员', '公益组织负责人', '内容创作者']),
            ('★★★☆☆ 中影响', ['产品经理', 'CEO/创业者', '管理咨询顾问', '品牌营销经理']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '算法工程师', '硬件工程师']),
        ]
    },
    {
        'id': 'transcend_04', 'name': '超越性信念（灵性感知）', 'category': '超越动力',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '战略规划师', '投资分析师', '行业研究员', '创新研究员']),
            ('★★★☆☆ 中影响', ['产品经理', '运营总监', '市场总监', '内容创作者']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '行政专员']),
        ]
    },
    {
        'id': 'transcend_05', 'name': '欣赏美与卓越', 'category': '超越动力',
        'jobs': [
            ('★★★★★ 强影响', ['UI设计师', 'UX设计师', '品牌设计师', '插画师', '视频剪辑师', '工业设计师']),
            ('★★★☆☆ 中影响', ['产品经理', '内容创作者', '品牌营销经理', '前端开发工程师']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '基础客服', '算法工程师', '审计专员']),
        ]
    },
    {
        'id': 'transcend_06', 'name': '超越性感恩与宽恕', 'category': '超越动力',
        'jobs': [
            ('★★★★★ 强影响', ['CEO/创业者', '运营总监', 'HRD/CHO', '销售经理', '客户成功经理', '培训发展专员']),
            ('★★★☆☆ 中影响', ['产品经理', '项目经理', 'HRBP', '管理咨询顾问', '行政专员']),
            ('☆☆☆☆☆ 无影响', ['流水线操作', '数据录入员', '硬件工程师', '算法工程师']),
        ]
    },
]

# 创建sheet
ws_main = wb.create_sheet("题目影响总表")
ws_main.sheet_view.showGridLines = False

# 合并大标题
ws_main['A1'] = '27个子维度 × 50个岗位 — 测评题目对工作的影响程度矩阵'
ws_main['A1'].font = Font(name='Arial', bold=True, size=13)
ws_main.merge_cells('A1:AX1')
ws_main['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_main.row_dimensions[1].height = 28

# 读取50个工作名称（从上表数据中提取）
all_jobs = [
    '后端开发工程师', '前端开发工程师', '算法工程师', '数据分析师', '硬件工程师', '测试工程师',
    'UI设计师', 'UX设计师', '品牌设计师', '工业设计师', '视频剪辑师', '插画师',
    '产品经理', '项目经理', '创新研究员', '内容创作者', '知识产权专员',
    '战略规划师', '管理咨询顾问', '投资分析师', '商业分析师', '行业研究员',
    '电商运营', '用户运营', '内容运营', '供应链专员', '质量管理专员',
    '品牌营销经理', '数字营销专员', '销售经理', '大客户经理', '客户成功经理', 'BD商务拓展',
    '财务分析师', '审计专员', '投融资专员', '税务专员',
    'HRBP', '招聘专员', '培训发展专员', 'OD组织发展专员',
    '行政专员', '法务专员', '数据录入员', '客服专员',
    '运营总监', '市场总监', '技术VP/CTO', 'HRD/CHO', 'CEO/创业者'
]

# 列宽
ws_main.column_dimensions['A'].width = 26
ws_main.column_dimensions['B'].width = 14
for i in range(3, 53):
    ws_main.column_dimensions[get_column_letter(i)].width = 12

# 写表头（第2行）
ws_main.cell(row=2, column=1, value='子维度名称').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws_main.cell(row=2, column=1).fill = PatternFill('solid', fgColor='2F5496')
ws_main.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_main.cell(row=2, column=2, value='所属大类').font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
ws_main.cell(row=2, column=2).fill = PatternFill('solid', fgColor='2F5496')
ws_main.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')

for col, job in enumerate(all_jobs, 3):
    c = ws_main.cell(row=2, column=col, value=job)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

ws_main.row_dimensions[2].height = 35

# 填数据
for row_idx, dim in enumerate(dimensions, 3):
    # 子维度名
    c = ws_main.cell(row=row_idx, column=1, value=dim['name'])
    c.font = Font(name='Arial', bold=True, size=9)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border = border

    # 大类
    c = ws_main.cell(row=row_idx, column=2, value=dim['category'])
    c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

    # 构建岗位影响字典
    job_influence = {}
    for level, jobs_list in dim['jobs']:
        for job in jobs_list:
            job_influence[job] = level

    # 逐列填影响力
    for col, job in enumerate(all_jobs, 3):
        influence = job_influence.get(job, '☆☆☆☆☆ 无影响')
        c = ws_main.cell(row=row_idx, column=col, value=influence)
        c.font = Font(name='Arial', size=8)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        if '★★★★★' in influence:
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        elif '★★★' in influence:
            c.fill = PatternFill('solid', fgColor='FFEB9C')
        else:
            c.fill = PatternFill('solid', fgColor='FFCCCC')

    ws_main.row_dimensions[row_idx].height = 30

# 添加图例行
legend_row = len(dimensions) + 4
ws_main.cell(row=legend_row, column=1, value='■ 图例').font = Font(name='Arial', bold=True, size=10)
ws_main.merge_cells(f'A{legend_row}:B{legend_row}')

legend_items = [
    ('★★★★★ 强影响', 'C6EFCE', '该维度高低直接决定能否胜任该工作的核心要求'),
    ('★★★☆☆ 中影响', 'FFEB9C', '该维度对工作有正向帮助，高分有助于提升工作表现'),
    ('☆☆☆☆☆ 无影响', 'FFCCCC', '该维度在该工作中不是关键能力，高低分均不影响'),
]
for i, (txt, color, desc) in enumerate(legend_items):
    r = legend_row + 1 + i
    ws_main.cell(row=r, column=1, value=txt).font = Font(name='Arial', size=9)
    ws_main.cell(row=r, column=1).fill = PatternFill('solid', fgColor=color)
    ws_main.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws_main.merge_cells(f'B{r}:D{r}')
    ws_main.cell(row=r, column=2, value=desc).font = Font(name='Arial', size=9)
    ws_main.cell(row=r, column=2).alignment = Alignment(vertical='center')

ws_main.freeze_panes = 'C3'

wb.save(r'C:\Users\25890\Desktop\题目影响分析表_个人成长名片.xlsx')
print('题目影响分析表_个人成长名片.xlsx 创建成功！')
