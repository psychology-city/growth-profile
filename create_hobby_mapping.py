# 兴趣爱好与子维度关系表 - 重写版
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取详细数据
with open(r'C:\Users\25890\.qclaw\workspace\growth-profile\detail_data.json', 'r', encoding='utf-8') as f:
    detail_raw = json.load(f)

wb = Workbook()
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============ Sheet 1: 总关系矩阵 ============
ws = wb.active
ws.title = "关系矩阵"
ws.sheet_view.showGridLines = False

ws['A1'] = '42个兴趣爱好 x 27个子维度 - 适配关系矩阵（成长型推荐视角）'
ws['A1'].font = Font(name='Arial', bold=True, size=12)
ws.merge_cells('A1:ABA1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

hobbies_short = [
    '跑步','健身','游泳','球类','骑行','格斗','瑜伽','极限',
    '乐器','唱歌','绘画','书法','摄影','舞蹈','手工艺',
    '阅读','写作','语言','口播',
    '游戏','桌游','棋类',
    '旅行','徒步','观鸟','钓鱼',
    '烹饪','品酒','园艺','收藏',
    '编程','AI工具','硬件','无人机',
    '志愿','社群','辩论','疗愈',
    '投资','历史','科学','考证'
]

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 12
for i in range(3, 45):
    ws.column_dimensions[get_column_letter(i)].width = 8

# 表头
for col, text in enumerate(['兴趣爱好', '大类'], 1):
    c = ws.cell(row=2, column=col, value=text)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

dims_short = ['注意力','思维洞察','好奇心','信息分析','元认知','知识迁移',
              '开放判断','创造力','实践落地','共情内核','关系构建',
              '协作沟通','赋能领导','内在信念','目标推进','自我管控',
              '决策引领','价值交付','前瞻领导','落地统筹','魅力调节',
              '自我超越','意义建构','利他贡献','超越信念','欣赏美','超越感恩']

for col, dim in enumerate(dims_short, 3):
    c = ws.cell(row=2, column=col, value=dim)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws.row_dimensions[2].height = 50

hobby_cats = [
    '运动健身','运动健身','运动健身','运动健身','运动健身','运动健身','运动健身','运动健身',
    '音乐艺术','音乐艺术','音乐艺术','音乐艺术','音乐艺术','音乐艺术','音乐艺术',
    '知识学习','知识学习','知识学习','知识学习',
    '游戏娱乐','游戏娱乐','游戏娱乐',
    '户外探索','户外探索','户外探索','户外探索',
    '生活美食','生活美食','生活美食','生活美食',
    '科技数码','科技数码','科技数码','科技数码',
    '社交公益','社交公益','社交公益','社交公益',
    '知识学习','知识学习','知识学习','知识学习'
]

# 关系矩阵 (1=核心匹配 0=无)
hobby_matrix = [
    [1,0,0,0,0,0,0,0,0,0,0,0,0,1,1,1,0,0,0,0,1,1,0,1,0,1],
    [0,0,1,0,0,0,0,0,1,0,0,0,1,1,1,1,1,0,1,1,1,1,0,0,0,1],
    [1,0,0,0,0,0,0,0,0,0,0,0,0,1,1,1,0,0,0,1,0,1,0,0,0,1],
    [0,0,0,0,0,0,0,0,1,1,1,1,1,0,1,0,1,0,1,1,1,0,0,0,0,1],
    [0,0,0,0,0,0,0,0,1,0,1,0,0,1,1,1,0,0,0,1,1,1,0,0,0,1],
    [0,0,0,0,0,0,0,0,1,1,1,1,1,1,1,1,1,0,1,1,1,0,0,1,0,1],
    [1,0,1,0,1,0,0,0,1,0,0,0,0,1,1,1,0,0,0,1,1,1,0,1,0,1],
    [0,0,0,0,0,0,0,0,1,0,0,0,0,1,1,1,1,0,1,1,1,0,0,1,0,0],
    [1,1,0,0,1,0,0,0,1,0,0,0,0,1,1,1,1,0,0,1,0,1,0,1,0,1],
    [1,1,0,0,1,0,0,0,0,1,1,1,0,0,0,0,0,0,0,0,1,0,0,0,1,1],
    [1,1,1,0,0,1,1,1,1,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,1,0],
    [1,0,1,0,1,1,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,1,0,1,0,1],
    [0,1,1,0,0,1,1,1,1,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,1,0],
    [0,0,0,0,0,0,1,1,1,1,1,1,1,0,0,0,0,0,1,1,1,0,0,0,1,0],
    [0,0,1,0,0,1,0,1,1,0,0,0,0,0,1,0,1,0,0,1,0,0,0,0,0,0],
    [0,1,1,0,1,1,1,0,0,0,0,0,0,1,0,0,0,0,0,0,0,1,0,1,0,1],
    [0,0,1,0,1,1,1,1,0,0,0,0,0,1,0,0,1,0,0,0,1,1,0,0,0,0],
    [0,0,1,0,1,1,1,0,0,1,0,0,0,1,1,0,0,0,0,0,0,1,1,0,0,0],
    [0,0,0,0,0,0,1,0,0,1,0,1,0,0,0,0,0,0,0,0,1,0,1,0,0,1],
    [1,1,0,1,0,0,1,0,1,1,1,0,1,0,1,0,1,0,1,0,1,0,0,0,0,1],
    [0,1,0,1,0,0,1,0,0,1,1,1,0,0,0,0,0,0,1,0,1,0,0,0,0,0],
    [1,1,0,1,1,0,0,0,0,0,0,0,0,0,1,1,0,0,0,0,0,0,0,0,0,0],
    [0,0,1,0,0,1,1,0,0,1,1,1,0,1,0,0,1,0,1,0,1,1,1,1,0,1],
    [0,0,1,0,0,0,0,0,1,1,1,1,1,1,1,0,1,0,1,1,1,1,1,1,0,1],
    [1,0,1,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [1,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,0,0,0,0,0,0,0,0,1],
    [0,0,1,0,0,1,0,1,1,1,0,0,0,0,1,0,0,1,0,0,0,0,0,0,1,0],
    [0,0,0,0,0,0,1,0,0,1,1,0,0,1,0,0,0,0,0,0,1,1,0,1,1,1],
    [1,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,1,0,0,0,1],
    [0,0,1,0,0,0,1,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,1,0],
    [1,1,1,0,1,0,0,0,1,0,0,0,0,0,1,0,0,0,0,1,0,0,0,0,0,0],
    [0,0,1,0,0,0,1,1,1,0,0,0,0,0,1,0,0,0,0,1,0,0,0,0,0,0],
    [1,0,1,0,0,0,0,0,1,0,0,0,0,0,1,0,0,0,0,1,0,0,0,0,0,0],
    [1,0,0,0,0,0,0,0,1,0,0,0,0,0,1,1,0,0,0,1,0,0,0,0,0,0],
    [0,0,0,0,0,0,0,0,0,1,1,1,1,1,1,0,1,1,1,1,1,0,1,0,0,1],
    [0,0,0,0,0,0,0,0,1,1,1,1,1,0,1,0,1,0,1,1,1,0,1,0,0,0],
    [0,1,0,0,0,0,1,0,0,1,0,1,0,0,0,0,0,0,1,0,1,0,1,0,0,0],
    [0,0,0,0,1,0,0,0,0,1,0,0,0,1,0,0,0,0,0,0,0,1,0,1,0,1],
    [0,1,0,1,1,0,0,0,0,0,0,0,0,0,1,1,0,0,0,0,0,0,0,1,0,0],
    [0,1,1,0,1,1,1,0,0,0,0,0,0,1,0,0,0,0,0,0,0,1,0,1,0,0],
    [1,1,1,0,1,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    [0,0,1,0,0,0,0,0,1,0,0,0,0,1,1,1,0,0,0,0,0,0,0,0,0,0],
]

for row_idx, (hobby, cat) in enumerate(zip(hobbies_short, hobby_cats), 3):
    c = ws.cell(row=row_idx, column=1, value=hobby)
    c.font = Font(name='Arial', bold=True, size=9)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = border

    c = ws.cell(row=row_idx, column=2, value=cat)
    c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

    row_matrix = hobby_matrix[row_idx - 3]
    for dim_idx, val in enumerate(row_matrix):
        c = ws.cell(row=row_idx, column=dim_idx + 3, value='●' if val == 1 else '○')
        c.font = Font(name='Arial', size=8, color='1F7A1F' if val == 1 else 'AAAAAA')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        if val == 1:
            c.fill = PatternFill('solid', fgColor='C6EFCE')
        else:
            c.fill = PatternFill('solid', fgColor='F5F5F5')

    ws.row_dimensions[row_idx].height = 22

# 图例
lr = len(hobby_matrix) + 5
ws.cell(row=lr, column=1, value='图例').font = Font(name='Arial', bold=True, size=10)
for i, (sym, color, desc) in enumerate([
    ('●', 'C6EFCE', '核心匹配：成长型模式高分强烈推荐该兴趣，能深度沉浸并持续精进'),
    ('○', 'F5F5F5', '无影响：该维度对该兴趣爱好基本没有影响'),
]):
    r = lr + 1 + i
    ws.cell(row=r, column=1, value=sym).font = Font(name='Arial', size=11, color='1F7A1F' if i == 0 else 'AAAAAA', bold=True)
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=2, value=desc).font = Font(name='Arial', size=10)
    ws.merge_cells(f'B{r}:G{r}')

ws.freeze_panes = 'C3'

# ============ Sheet 2: 详细推荐 ============
ws2 = wb.create_sheet("推荐理由")
ws2.sheet_view.showGridLines = False
ws2['A1'] = '兴趣爱好 - 适配推荐与成长建议'
ws2['A1'].font = Font(name='Arial', bold=True, size=13)
ws2.merge_cells('A1:F1')
ws2.row_dimensions[1].height = 28

headers2 = ['兴趣爱好', '大类', '核心匹配维度(成长型高分)', '延伸匹配维度', '成长建议', '等级型高分者适配方向']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws2.row_dimensions[2].height = 30

col_widths2 = [14, 10, 38, 30, 38, 35]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for row_idx, row_data in enumerate(detail_raw, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = border
    ws2.row_dimensions[row_idx].height = 70

ws2.freeze_panes = 'A3'

output_path = r'C:\Users\25890\Desktop\兴趣爱好\兴趣爱好与子维度关系.xlsx'
wb.save(output_path)
print(f'Successfully saved to: {output_path}')
